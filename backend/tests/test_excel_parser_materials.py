from pathlib import Path

from openpyxl import Workbook

from app.services.excel_parser import ExcelEstimateParser


def _save_workbook(tmp_path: Path, name: str, fill_rows) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = name
    fill_rows(ws)
    path = tmp_path / f"{name}.xlsx"
    wb.save(path)
    wb.close()
    return path


def test_structured_type2_parser_keeps_materials_under_work(tmp_path: Path) -> None:
    def fill(ws):
        headers = [
            "",
            "ВИД РАБОТЫ",
            "МАТЕРИАЛЫ",
            "Ед.изм",
            "Ед.изм",
            "Кол-во",
            "Кол-во",
            "Цена",
            "Цена с наценкой",
            "Цена",
            "Сумма",
            "Сумма",
            "ИТОГО",
        ]
        for idx, value in enumerate(headers, start=1):
            ws.cell(4, idx).value = value

        ws.cell(5, 2).value = "ПОЛЫ"
        ws.cell(5, 2).font = ws.cell(5, 2).font.copy(bold=True)

        ws.cell(6, 2).value = "Устройство стяжки"
        ws.cell(6, 4).value = "м2"
        ws.cell(6, 6).value = 18
        ws.cell(6, 8).value = 450
        ws.cell(6, 11).value = 8100

        ws.cell(7, 3).value = "Цемент М500"
        ws.cell(7, 5).value = "меш"
        ws.cell(7, 7).value = 9
        ws.cell(7, 10).value = 380
        ws.cell(7, 12).value = 3420

        ws.cell(8, 3).value = "Песок"
        ws.cell(8, 5).value = "м3"
        ws.cell(8, 7).value = 1.2
        ws.cell(8, 10).value = 950
        ws.cell(8, 12).value = 1140

    path = _save_workbook(tmp_path, "Смета", fill)

    rows, meta = ExcelEstimateParser().parse(path)

    assert meta["strategy"] == "structured_smeta"
    assert len(rows) == 1
    assert rows[0].section == "ПОЛЫ"
    assert rows[0].work_name == "Устройство стяжки"
    assert rows[0].quantity == 18
    assert rows[0].total_price == 8100
    assert len(rows[0].materials) == 2
    assert rows[0].materials[0]["name"] == "Цемент М500"
    assert rows[0].materials[0]["quantity"] == 9
    assert rows[0].materials[1]["name"] == "Песок"
    assert rows[0].materials[1]["total_price"] == 1140


def test_structured_type2_uses_rightmost_total_column(tmp_path: Path) -> None:
    def fill(ws):
        ws.title = "СМЕТА"
        headers = [
            "",
            "ВИД РАБОТЫ",
            "ИСПОЛЬЗУЕМЫЕ МАТЕРИАЛЫ",
            "Ед.изм. Работ",
            "Ед.изм. Материала",
            "Кол-во работ",
            "Кол-во материала",
            "Цена ед. Работ",
            "Цена ед. Работ",
            "Цена ед. Материала",
            "Сумма работ",
            "Сумма материалов",
            "ИТОГО",
            "Стоимость ед. работ и  материалов",
            "Сумма работ  и материалов",
        ]
        for idx, value in enumerate(headers, start=1):
            ws.cell(4, idx).value = value

        ws.cell(5, 2).value = "ПОЛЫ"

        ws.cell(6, 2).value = "Устройство пола"
        ws.cell(6, 4).value = "м2"
        ws.cell(6, 6).value = 10
        ws.cell(6, 8).value = 100
        ws.cell(6, 11).value = 1000
        ws.cell(6, 13).value = 1000
        ws.cell(6, 15).value = 1450

        ws.cell(7, 3).value = "Смесь"
        ws.cell(7, 5).value = "меш"
        ws.cell(7, 7).value = 5
        ws.cell(7, 10).value = 90
        ws.cell(7, 12).value = 450
        ws.cell(7, 13).value = 450

    path = _save_workbook(tmp_path, "ОбразецИтого", fill)

    rows, meta = ExcelEstimateParser().parse(path)

    assert meta["strategy"] == "structured_smeta"
    assert len(rows) == 1
    assert rows[0].work_name == "Устройство пола"
    assert rows[0].total_price == 1450
    assert rows[0].materials[0]["total_price"] == 450


def test_structured_type1_parser_attaches_material_rows_to_previous_work(tmp_path: Path) -> None:
    def fill(ws):
        ws.title = "Журнал"
        headers = ["№", "Позиция", "Тип", "Ед.изм", "Кол-во", "Цена за ед.", "Стоимость"]
        for idx, value in enumerate(headers, start=1):
            ws.cell(1, idx).value = value

        ws.cell(2, 1).value = 1
        ws.cell(2, 2).value = "Внутренняя отделка"

        ws.cell(3, 1).value = "1.1"
        ws.cell(3, 2).value = "Штукатурка стен"
        ws.cell(3, 3).value = "Работа"
        ws.cell(3, 4).value = "м2"
        ws.cell(3, 5).value = 32
        ws.cell(3, 6).value = 520
        ws.cell(3, 7).value = 16640

        ws.cell(4, 1).value = "1.2"
        ws.cell(4, 2).value = "Штукатурная смесь"
        ws.cell(4, 3).value = "Материал"
        ws.cell(4, 4).value = "меш"
        ws.cell(4, 5).value = 16
        ws.cell(4, 6).value = 340
        ws.cell(4, 7).value = 5440

        ws.cell(5, 1).value = "1.3"
        ws.cell(5, 2).value = "Доставка смеси"
        ws.cell(5, 3).value = "Накладные"
        ws.cell(5, 4).value = "рейс"
        ws.cell(5, 5).value = 1
        ws.cell(5, 6).value = 1800
        ws.cell(5, 7).value = 1800

    path = _save_workbook(tmp_path, "Лист1", fill)

    rows, meta = ExcelEstimateParser().parse(path)

    assert meta["strategy"] == "structured_smeta"
    assert len(rows) == 2
    assert rows[0].section == "Внутренняя отделка"
    assert rows[0].work_name == "Штукатурка стен"
    assert rows[0].total_price == 16640
    assert rows[0].materials == [
        {
            "name": "Штукатурная смесь",
            "unit": "меш",
            "quantity": 16.0,
            "unit_price": 340.0,
            "total_price": 5440.0,
        }
    ]
    assert rows[1].work_name == "Доставка смеси"
    assert rows[1].unit_price == 1800
    assert rows[1].total_price == 1800
