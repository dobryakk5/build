from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl, unquote

import asyncpg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BACKEND_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = BACKEND_DIR.parent

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

DEFAULT_OUTPUT_DIR = PROJECT_DIR / "output" / "fer_collections_xlsx"


SECTION_HEADERS = [
    "section_id",
    "collection_id",
    "collection_num",
    "collection_name",
    "section_title",
    "ignored",
    "effective_ignored",
    "subsection_count",
    "direct_table_count",
    "total_table_count",
]

SUBSECTION_HEADERS = [
    "subsection_id",
    "section_id",
    "collection_id",
    "collection_num",
    "collection_name",
    "section_title",
    "subsection_title",
    "ignored",
    "effective_ignored",
    "table_count",
]

TABLE_HEADERS = [
    "table_id",
    "collection_id",
    "collection_num",
    "collection_name",
    "section_id",
    "section_title",
    "subsection_id",
    "subsection_title",
    "placement",
    "table_title",
    "common_work_name",
    "row_count",
    "table_url",
    "ignored",
    "effective_ignored",
    "scraped_at",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export FER hierarchy into one XLSX file per collection. "
            "Each file contains fer.sections, fer.subsections, and fer.fer_tables sheets."
        )
    )
    parser.add_argument(
        "--database-url",
        default=None,
        help="Database URL. Defaults to DATABASE_URL from backend/.env or environment.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help=f"Directory for XLSX files. Default: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--expected-collections",
        type=int,
        default=47,
        help="Expected collection count; prints a warning if the database has another count.",
    )
    parser.add_argument(
        "--collection-num",
        action="append",
        default=[],
        help="Export only selected collection number. Can be passed multiple times, e.g. --collection-num 01.",
    )
    parser.add_argument(
        "--max-tables",
        type=int,
        default=None,
        help="Test export limit: include only the first N fer.fer_tables rows across selected collections.",
    )
    return parser.parse_args()


def load_database_url(cli_value: str | None) -> str:
    if cli_value:
        database_url = cli_value
    else:
        load_dotenv(BACKEND_DIR / ".env", override=False)
        database_url = os.getenv("DATABASE_URL", "")

    if not database_url:
        raise RuntimeError("DATABASE_URL is not set. Pass --database-url or configure backend/.env.")

    if database_url.startswith("postgresql+asyncpg://"):
        database_url = database_url.replace("postgresql+asyncpg://", "postgresql://", 1)

    return database_url


def parse_database_url(database_url: str) -> dict[str, Any]:
    if "://" not in database_url:
        raise ValueError("DATABASE_URL must include a scheme, e.g. postgresql://")

    scheme, rest = database_url.split("://", 1)
    if scheme not in {"postgresql", "postgres"}:
        raise ValueError(f"Unsupported database scheme: {scheme}")
    if "@" not in rest:
        raise ValueError("DATABASE_URL must include user credentials and host.")

    userinfo, host_and_database = rest.rsplit("@", 1)
    if ":" in userinfo:
        user, password = userinfo.split(":", 1)
    else:
        user, password = userinfo, ""

    if "/" in host_and_database:
        host_port, database_and_query = host_and_database.split("/", 1)
    else:
        host_port, database_and_query = host_and_database, ""

    database, _, query = database_and_query.partition("?")

    if host_port.startswith("["):
        host, _, port_part = host_port[1:].partition("]")
        port = port_part[1:] if port_part.startswith(":") else ""
    elif ":" in host_port:
        host, port = host_port.rsplit(":", 1)
    else:
        host, port = host_port, ""

    kwargs: dict[str, Any] = {
        "user": unquote(user),
        "password": unquote(password),
        "host": host,
        "database": unquote(database),
    }
    if port:
        kwargs["port"] = int(port)

    for key, value in parse_qsl(query, keep_blank_values=True, strict_parsing=False):
        if key == "sslmode":
            if value in {"require", "verify-ca", "verify-full"}:
                kwargs["ssl"] = True
        elif key in {"application_name", "server_settings"}:
            continue

    return kwargs


async def connect_database(database_url: str) -> asyncpg.Connection:
    return await asyncpg.connect(**parse_database_url(database_url))


def safe_filename(value: str) -> str:
    value = value.replace("№", "No")
    value = re.sub(r'[\\/*?:"<>|]+', "_", value)
    value = re.sub(r"\s+", " ", value).strip()
    value = value.rstrip(". ")
    return value[:160] or "collection"


def normalize_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def rows_to_values(rows: Iterable[asyncpg.Record], headers: list[str]) -> list[list[Any]]:
    return [[normalize_value(row[header]) for header in headers] for row in rows]


def autofit_columns(ws) -> None:
    max_width = 72
    min_width = 10

    for column_cells in ws.columns:
        column_index = column_cells[0].column
        letter = get_column_letter(column_index)
        longest = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


async def fetch_collections(
    conn: asyncpg.Connection,
    collection_nums: list[str],
    max_tables: int | None,
) -> list[asyncpg.Record]:
    if max_tables is not None:
        sql = """
            WITH limited_tables AS (
                SELECT t.collection_id
                FROM fer.fer_tables t
                JOIN fer.collections c ON c.id = t.collection_id
                WHERE CARDINALITY($1::text[]) = 0 OR c.num = ANY($1::text[])
                ORDER BY c.num, t.id
                LIMIT $2
            )
            SELECT DISTINCT
                c.id,
                c.num,
                c.name,
                COALESCE(c.ignored, FALSE) AS ignored
            FROM fer.collections c
            JOIN limited_tables lt ON lt.collection_id = c.id
            ORDER BY c.num
        """
        return await conn.fetch(sql, collection_nums, max_tables)

    sql = """
        SELECT
            c.id,
            c.num,
            c.name,
            COALESCE(c.ignored, FALSE) AS ignored
        FROM fer.collections c
        WHERE CARDINALITY($1::text[]) = 0 OR c.num = ANY($1::text[])
        ORDER BY c.num
    """
    return await conn.fetch(sql, collection_nums)


async def fetch_sections(
    conn: asyncpg.Connection,
    collection_id: int,
    table_ids: list[int] | None,
) -> list[asyncpg.Record]:
    return await conn.fetch(
        """
        WITH scoped_tables AS (
            SELECT *
            FROM fer.fer_tables
            WHERE CARDINALITY($2::int[]) = 0 OR id = ANY($2::int[])
        ),
        scoped_sections AS (
            SELECT DISTINCT s.id
            FROM fer.sections s
            LEFT JOIN scoped_tables t ON t.section_id = s.id
            WHERE s.collection_id = $1
              AND (CARDINALITY($2::int[]) = 0 OR t.id IS NOT NULL)
        )
        SELECT
            s.id AS section_id,
            s.collection_id,
            c.num AS collection_num,
            c.name AS collection_name,
            s.title AS section_title,
            COALESCE(s.ignored, FALSE) AS ignored,
            (COALESCE(c.ignored, FALSE) OR COALESCE(s.ignored, FALSE)) AS effective_ignored,
            COUNT(DISTINCT ss.id)::int AS subsection_count,
            COUNT(DISTINCT direct_t.id)::int AS direct_table_count,
            COUNT(DISTINCT all_t.id)::int AS total_table_count
        FROM fer.sections s
        JOIN scoped_sections scoped_s ON scoped_s.id = s.id
        JOIN fer.collections c ON c.id = s.collection_id
        LEFT JOIN fer.subsections ss ON ss.section_id = s.id
        LEFT JOIN scoped_tables direct_t
            ON direct_t.section_id = s.id
           AND direct_t.subsection_id IS NULL
        LEFT JOIN scoped_tables all_t ON all_t.section_id = s.id
        WHERE s.collection_id = $1
        GROUP BY s.id, s.collection_id, c.num, c.name, s.title, s.ignored, c.ignored
        ORDER BY s.id
        """,
        collection_id,
        table_ids or [],
    )


async def fetch_subsections(
    conn: asyncpg.Connection,
    collection_id: int,
    table_ids: list[int] | None,
) -> list[asyncpg.Record]:
    return await conn.fetch(
        """
        WITH scoped_tables AS (
            SELECT *
            FROM fer.fer_tables
            WHERE CARDINALITY($2::int[]) = 0 OR id = ANY($2::int[])
        )
        SELECT
            ss.id AS subsection_id,
            ss.section_id,
            s.collection_id,
            c.num AS collection_num,
            c.name AS collection_name,
            s.title AS section_title,
            ss.title AS subsection_title,
            COALESCE(ss.ignored, FALSE) AS ignored,
            (
                COALESCE(c.ignored, FALSE)
                OR COALESCE(s.ignored, FALSE)
                OR COALESCE(ss.ignored, FALSE)
            ) AS effective_ignored,
            COUNT(DISTINCT t.id)::int AS table_count
        FROM fer.subsections ss
        JOIN fer.sections s ON s.id = ss.section_id
        JOIN fer.collections c ON c.id = s.collection_id
        LEFT JOIN scoped_tables t ON t.subsection_id = ss.id
        WHERE s.collection_id = $1
          AND (CARDINALITY($2::int[]) = 0 OR t.id IS NOT NULL)
        GROUP BY
            ss.id,
            ss.section_id,
            s.collection_id,
            c.num,
            c.name,
            s.title,
            ss.title,
            ss.ignored,
            s.ignored,
            c.ignored
        ORDER BY ss.section_id, ss.id
        """,
        collection_id,
        table_ids or [],
    )


async def fetch_tables(
    conn: asyncpg.Connection,
    collection_id: int,
    table_ids: list[int] | None,
) -> list[asyncpg.Record]:
    return await conn.fetch(
        """
        SELECT
            t.id AS table_id,
            t.collection_id,
            c.num AS collection_num,
            c.name AS collection_name,
            t.section_id,
            s.title AS section_title,
            t.subsection_id,
            ss.title AS subsection_title,
            CASE
                WHEN t.section_id IS NULL THEN 'collection'
                WHEN t.subsection_id IS NULL THEN 'section'
                ELSE 'subsection'
            END AS placement,
            t.table_title,
            t.common_work_name,
            t.row_count::int AS row_count,
            t.table_url,
            COALESCE(t.ignored, FALSE) AS ignored,
            (
                COALESCE(c.ignored, FALSE)
                OR COALESCE(s.ignored, FALSE)
                OR COALESCE(ss.ignored, FALSE)
                OR COALESCE(t.ignored, FALSE)
            ) AS effective_ignored,
            t.scraped_at
        FROM fer.fer_tables t
        JOIN fer.collections c ON c.id = t.collection_id
        LEFT JOIN fer.sections s ON s.id = t.section_id
        LEFT JOIN fer.subsections ss ON ss.id = t.subsection_id
        WHERE t.collection_id = $1
          AND (CARDINALITY($2::int[]) = 0 OR t.id = ANY($2::int[]))
        ORDER BY
            COALESCE(t.section_id, 0),
            COALESCE(t.subsection_id, 0),
            t.id
        """,
        collection_id,
        table_ids or [],
    )


async def fetch_limited_table_ids(
    conn: asyncpg.Connection,
    collection_nums: list[str],
    max_tables: int | None,
) -> list[int] | None:
    if max_tables is None:
        return None
    if max_tables < 1:
        raise ValueError("--max-tables must be greater than 0")

    rows = await conn.fetch(
        """
        SELECT t.id
        FROM fer.fer_tables t
        JOIN fer.collections c ON c.id = t.collection_id
        WHERE CARDINALITY($1::text[]) = 0 OR c.num = ANY($1::text[])
        ORDER BY c.num, t.id
        LIMIT $2
        """,
        collection_nums,
        max_tables,
    )
    return [int(row["id"]) for row in rows]


async def export_collection(
    conn: asyncpg.Connection,
    collection: asyncpg.Record,
    output_dir: Path,
    table_ids: list[int] | None,
) -> Path:
    sections = await fetch_sections(conn, int(collection["id"]), table_ids)
    subsections = await fetch_subsections(conn, int(collection["id"]), table_ids)
    tables = await fetch_tables(conn, int(collection["id"]), table_ids)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_sheet(wb, "fer.sections", SECTION_HEADERS, rows_to_values(sections, SECTION_HEADERS))
    write_sheet(wb, "fer.subsections", SUBSECTION_HEADERS, rows_to_values(subsections, SUBSECTION_HEADERS))
    write_sheet(wb, "fer.fer_tables", TABLE_HEADERS, rows_to_values(tables, TABLE_HEADERS))

    output_dir.mkdir(parents=True, exist_ok=True)
    collection_num = str(collection["num"])
    collection_name = str(collection["name"])
    filename = safe_filename(f"FER_{collection_num}_{collection_name}.xlsx")
    path = output_dir / filename
    wb.save(path)
    return path


async def main() -> None:
    args = parse_args()
    database_url = load_database_url(args.database_url)
    output_dir = Path(args.output_dir).expanduser().resolve()

    conn = await connect_database(database_url)
    try:
        table_ids = await fetch_limited_table_ids(conn, args.collection_num, args.max_tables)
        collections = await fetch_collections(conn, args.collection_num, args.max_tables)
        if (
            args.expected_collections
            and not args.collection_num
            and args.max_tables is None
            and len(collections) != args.expected_collections
        ):
            print(
                f"[WARN] Expected {args.expected_collections} collections, "
                f"but found {len(collections)}.",
                file=sys.stderr,
            )

        if not collections:
            print("No FER collections found.")
            return

        limit_note = f" with {len(table_ids)} tables" if table_ids is not None else ""
        print(f"Exporting {len(collections)} FER collections{limit_note} to {output_dir}")
        for index, collection in enumerate(collections, start=1):
            path = await export_collection(conn, collection, output_dir, table_ids)
            print(f"[{index}/{len(collections)}] {collection['num']} -> {path}")
    finally:
        await conn.close()


if __name__ == "__main__":
    asyncio.run(main())
