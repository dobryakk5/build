from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, unquote

import asyncpg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BACKEND_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = BACKEND_DIR.parent
DEFAULT_OUTPUT = PROJECT_DIR / "output" / "fer_sections_flat.xlsx"


HEADERS = ["collection_num", "collection_name", "section_title"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export all FER collection sections into one XLSX sheet."
    )
    parser.add_argument(
        "--database-url",
        default=None,
        help="Database URL. Defaults to DATABASE_URL from backend/.env or environment.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help=f"Output XLSX path. Default: {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def load_database_url(cli_value: str | None) -> str:
    if cli_value:
        database_url = cli_value
    else:
        load_dotenv(BACKEND_DIR / ".env", override=False)
        database_url = os.getenv("DATABASE_URL", "")

    if not database_url:
        raise RuntimeError("DATABASE_URL is not set. Pass --database-url or configure backend/.env.")

    if database_url.startswith("postgresql+asyncpg://"):
        database_url = database_url.replace("postgresql+asyncpg://", "postgresql://", 1)

    return database_url


def parse_database_url(database_url: str) -> dict[str, Any]:
    if "://" not in database_url:
        raise ValueError("DATABASE_URL must include a scheme, e.g. postgresql://")

    scheme, rest = database_url.split("://", 1)
    if scheme not in {"postgresql", "postgres"}:
        raise ValueError(f"Unsupported database scheme: {scheme}")
    if "@" not in rest:
        raise ValueError("DATABASE_URL must include user credentials and host.")

    userinfo, host_and_database = rest.rsplit("@", 1)
    if ":" in userinfo:
        user, password = userinfo.split(":", 1)
    else:
        user, password = userinfo, ""

    if "/" in host_and_database:
        host_port, database_and_query = host_and_database.split("/", 1)
    else:
        host_port, database_and_query = host_and_database, ""

    database, _, query = database_and_query.partition("?")

    if host_port.startswith("["):
        host, _, port_part = host_port[1:].partition("]")
        port = port_part[1:] if port_part.startswith(":") else ""
    elif ":" in host_port:
        host, port = host_port.rsplit(":", 1)
    else:
        host, port = host_port, ""

    kwargs: dict[str, Any] = {
        "user": unquote(user),
        "password": unquote(password),
        "host": host,
        "database": unquote(database),
    }
    if port:
        kwargs["port"] = int(port)

    for key, value in parse_qsl(query, keep_blank_values=True, strict_parsing=False):
        if key == "sslmode" and value in {"require", "verify-ca", "verify-full"}:
            kwargs["ssl"] = True

    return kwargs


async def connect_database(database_url: str) -> asyncpg.Connection:
    return await asyncpg.connect(**parse_database_url(database_url))


async def fetch_rows(conn: asyncpg.Connection) -> list[asyncpg.Record]:
    return await conn.fetch(
        """
        SELECT
            c.num AS collection_num,
            c.name AS collection_name,
            s.title AS section_title
        FROM fer.collections c
        LEFT JOIN fer.sections s ON s.collection_id = c.id
        ORDER BY c.num, s.id NULLS FIRST
        """
    )


def autofit_columns(ws) -> None:
    for column_cells in ws.columns:
        column_index = column_cells[0].column
        letter = get_column_letter(column_index)
        longest = 0
        for cell in column_cells:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(14, min(80, longest + 2))


def write_workbook(rows: list[asyncpg.Record], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "fer.sections"
    ws.append(HEADERS)

    for row in rows:
        ws.append([row["collection_num"], row["collection_name"], row["section_title"]])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


async def main() -> None:
    args = parse_args()
    output = Path(args.output).expanduser().resolve()
    conn = await connect_database(load_database_url(args.database_url))
    try:
        rows = await fetch_rows(conn)
    finally:
        await conn.close()

    write_workbook(rows, output)
    print(f"Exported {len(rows)} rows to {output}")


if __name__ == "__main__":
    asyncio.run(main())
