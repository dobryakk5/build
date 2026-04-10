# backend/app/services/excel_parser.py
"""
Гибкий парсер Excel-смет.

Поддерживает два формата:
  1. СТРОЧНЫЙ  — классическая таблица с заголовком сверху (ГрандСмета, CourtDoc, Excel вручную)
  2. СТОЛБЦОВЫЙ — каждая работа = столбец, строки = атрибуты (редко, но встречается)

Алгоритм выбора стратегии:
  DetectorEngine сначала сканирует файл и выбирает стратегию,
  затем соответствующий Parser разбирает данные.

Если уверенность < 0.8 — поднимается NeedsMappingError с превью строк,
чтобы фронт показал UI ручного маппинга колонок.
"""

from __future__ import annotations
import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Protocol

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


# ─────────────────────────────────────────────────────────────────────────────
# DATA CLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ParsedRow:
    id: str = field(default_factory=lambda: str(uuid.uuid4()))
    section: Optional[str] = None
    work_name: str = ""
    unit: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None
    total_price: Optional[float] = None
    materials: list[dict] = field(default_factory=list)
    row_order: int = 0
    raw_data: dict = field(default_factory=dict)
    source_strategy: str = "unknown"   # для отладки: "row" | "column"


# ─────────────────────────────────────────────────────────────────────────────
# ОШИБКА: нужен ручной маппинг колонок
# ─────────────────────────────────────────────────────────────────────────────

class NeedsMappingError(Exception):
    """
    Поднимается когда парсер не смог уверенно определить колонки.
    Содержит превью данных, достаточное для показа UI маппинга клиенту.
    """
    def __init__(self, filename: str, sheet: str, preview_rows: list[list], col_count: int):
        self.filename     = filename
        self.sheet        = sheet
        self.preview_rows = preview_rows   # первые N строк сырых данных (без пустых)
        self.col_count    = col_count      # количество колонок
        super().__init__(f"Не удалось автоматически определить колонки в файле «{filename}»")


# ─────────────────────────────────────────────────────────────────────────────
# ALIAS DICTIONARY  (расширяемый, регистронезависимый)
# ─────────────────────────────────────────────────────────────────────────────

ALIASES: dict[str, list[str]] = {
    # ── Наименование работы ──────────────────────────────────────────────────
    "work_name": [
        # Русский
        "наименование", "наименование работ", "наименование работ и затрат",
        "наименование позиции", "описание работ", "вид работ",
        "вид работ и затрат", "работы", "работа", "содержание работ",
        "состав работ", "операция", "позиция", "статья затрат",
        "номенклатура", "наим.", "работы/материалы",
        # English
        "name", "work name", "description", "task", "activity",
        "item", "work item", "scope of work",
    ],

    # ── Единица измерения ────────────────────────────────────────────────────
    "unit": [
        # Русский
        "ед.изм", "ед. изм", "ед. изм.", "единица", "единица измерения",
        "единицы измерения", "ед", "ед.", "изм.", "измерение",
        "ед.изм.", "е.и.", "е. и.",
        # English
        "unit", "units", "uom", "u/m", "measure",
    ],

    # ── Количество / объём ───────────────────────────────────────────────────
    "quantity": [
        # Русский
        "кол-во", "кол.", "кол", "количество", "объем", "объём",
        "объем работ", "объём работ", "кол-во работ", "количество работ",
        "кол-во ед", "кол. ед.", "кол-во единиц", "итого объем",
        "всего объем", "объем по проекту", "проектный объем",
        "факт", "факт.кол", "фактическое количество",
        # English
        "qty", "quantity", "amount", "vol", "volume", "count",
    ],

    # ── Цена за единицу ──────────────────────────────────────────────────────
    "unit_price": [
        # Русский
        "цена", "цена за ед", "цена за ед.", "цена за единицу",
        "расценка", "стоимость ед", "стоимость ед.", "стоимость единицы",
        "цена ед.изм", "цена за ед.изм", "ед. стоимость",
        "стоим. ед.", "норм. цена", "базовая цена",
        "цена с ндс", "цена без ндс", "цена (без ндс)",
        # English
        "unit price", "price", "rate", "unit cost", "unit rate",
        "cost per unit", "price per unit",
    ],

    # ── Итоговая стоимость ───────────────────────────────────────────────────
    "total_price": [
        # Русский
        "сумма", "сумма всего", "итого", "итого стоимость",
        "общая стоимость", "стоимость", "стоимость работ",
        "стоимость всего", "стоимость итого", "итоговая стоимость",
        "всего", "всего сумма", "всего стоимость",
        "итог", "итог сумма", "итоговая сумма",
        "сумма с ндс", "сумма без ндс", "итого с ндс", "итого без ндс",
        "стоимость с ндс", "стоимость без ндс",
        "общая сумма", "общ. стоимость",
        # English
        "total", "total price", "total cost", "amount",
        "sum", "subtotal", "ext price", "extended price",
    ],

    # ── Трудоёмкость ─────────────────────────────────────────────────────────
    "labor_hours": [
        "трудоёмкость", "трудоемкость", "чел.-час", "чел.час", "чел/час",
        "чел-час", "норма времени", "затраты труда", "трудозатраты",
        "человеко-часы", "нормо-часы", "н/час", "н.час",
        "labor hours", "man hours", "manhours", "hours",
    ],

    # ── Код позиции (ЕНиР / ГЭСН) ────────────────────────────────────────────
    "enir_code": [
        "шифр", "код", "код расценки", "расценка", "норм. шифр",
        "шифр расценки", "ценник", "гэсн", "фер", "тер", "enir",
        "обоснование", "основание", "ссылка", "норматив",
        "code", "item code", "ref", "reference",
    ],

    # ── Материалы ─────────────────────────────────────────────────────────────
    "material_cost": [
        "материалы", "стоимость материалов", "мат.", "мат-лы",
        "материальные затраты", "стоимость матер",
        "materials", "material cost",
    ],
}


def normalize(s: str) -> str:
    """Нижний регистр, убрать лишние пробелы и спецсимволы для сравнения."""
    return re.sub(r"\s+", " ", str(s).lower().strip().replace("\n", " "))


def match_alias(cell_value: str, field_name: str) -> bool:
    """Проверяем, соответствует ли значение ячейки полю field_name."""
    val = normalize(cell_value)
    return any(alias in val or val in alias for alias in ALIASES[field_name])


def match_any_field(cell_value: str) -> Optional[str]:
    """Определяем поле по значению ячейки."""
    for field_name in ALIASES:
        if match_alias(cell_value, field_name):
            return field_name
    return None


_TYPE1_ITEM_RE = re.compile(r"^\d+\.\d+$")


def _cell_text(value) -> str:
    return str(value).strip() if value is not None else ""


def _to_optional_float(value, *, treat_zero_as_none: bool = False) -> Optional[float]:
    num = _to_float(value)
    if num is None:
        return None
    if treat_zero_as_none and num == 0:
        return None
    return num


def _is_empty_metric(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    num = _to_float(value)
    return num == 0 if num is not None else False


def _looks_like_group_label(text: str) -> bool:
    cleaned = re.sub(r"\s+", " ", text).strip()
    if not cleaned or len(cleaned) > 120:
        return False
    letters = re.sub(r"[^A-Za-zА-Яа-яЁё]", "", cleaned)
    return bool(letters) and (cleaned.isupper() or len(cleaned.split()) <= 8)


def _cell_is_emphasized(ws: Worksheet, row_idx: int, col_idx: int) -> bool:
    cell = ws.cell(row_idx, col_idx)
    font = getattr(cell, "font", None)
    if font and font.bold:
        return True

    fill = getattr(cell, "fill", None)
    if not fill or not getattr(fill, "fill_type", None) or fill.fill_type == "none":
        return False

    fg = getattr(fill, "fgColor", None)
    rgb = getattr(fg, "rgb", None) or getattr(fg, "value", None)
    if rgb is None:
        return fill.fill_type not in (None, "none")
    return str(rgb).upper() not in {"00000000", "000000", "00FFFFFF", "FFFFFFFF"}


def _make_material(
    name: str,
    *,
    unit=None,
    qty=None,
    price=None,
    total=None,
) -> dict:
    return {
        "name": name,
        "unit": _to_str(unit),
        "quantity": _to_optional_float(qty),
        "unit_price": _to_optional_float(price, treat_zero_as_none=True),
        "total_price": _to_optional_float(total, treat_zero_as_none=True),
    }


def _pick_rightmost_numeric(ws: Worksheet, row_idx: int, columns: list[int]) -> Optional[float]:
    for col_idx in sorted(set(columns), reverse=True):
        value = _to_optional_float(ws.cell(row_idx, col_idx).value, treat_zero_as_none=False)
        if value is not None:
            return value
    return None


# ─────────────────────────────────────────────────────────────────────────────
# СТРАТЕГИЯ 0: СТРУКТУРИРОВАННАЯ СМЕТА (группа → работа → материалы)
# ─────────────────────────────────────────────────────────────────────────────

class StructuredSmetaParser:
    """
    Приоритетная стратегия для смет вида:
      - № | Позиция | Тип | Ед.изм | Кол-во | Цена | Стоимость
      - ВИД РАБОТЫ | МАТЕРИАЛЫ | ...

    В БД сохраняются только строки работ, а материалы прикрепляются к ним
    в поле `materials`, чтобы не превращать материалы в задачи Ганта.
    """

    name = "structured_smeta"

    def can_parse(self, ws: Worksheet) -> float:
        if self._find_type1_header(ws):
            return 1.0
        if self._find_type2_header(ws):
            return 1.0
        return 0.0

    def parse(self, ws: Worksheet) -> list[ParsedRow]:
        if info := self._find_type1_header(ws):
            return self._parse_type1(ws, info)
        if info := self._find_type2_header(ws):
            return self._parse_type2(ws, info)
        return []

    def _find_type1_header(self, ws: Worksheet) -> Optional[dict]:
        for row_idx in range(1, min(ws.max_row + 1, 16)):
            mapping: dict[str, int] = {}
            price_cols: list[int] = []
            total_cols: list[int] = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                val = normalize(_cell_text(ws.cell(row_idx, col_idx).value))
                if not val:
                    continue
                if val in {"№", "номер"} or val.startswith("№"):
                    mapping.setdefault("num", col_idx)
                elif "тип" in val:
                    mapping.setdefault("type", col_idx)
                elif any(token in val for token in ("позиц", "наименование", "вид работ", "работы/материалы")):
                    mapping.setdefault("name", col_idx)
                elif "цен" in val:
                    price_cols.append(col_idx)
                elif "ед" in val:
                    mapping.setdefault("unit", col_idx)
                elif "кол" in val or "объем" in val or "объём" in val:
                    mapping.setdefault("qty", col_idx)
                elif any(token in val for token in ("стоим", "сумм", "итого", "всего")):
                    total_cols.append(col_idx)
            if {"num", "name", "type"} <= mapping.keys():
                if price_cols:
                    mapping["price"] = min(price_cols)
                if total_cols:
                    price_anchor = mapping.get("price", 0)
                    mapping["total"] = max([c for c in total_cols if c > price_anchor] or total_cols)
                mapping["header_row"] = row_idx
                return mapping
        return None

    def _find_type2_header(self, ws: Worksheet) -> Optional[dict]:
        for row_idx in range(1, min(ws.max_row + 1, 16)):
            work_col = material_col = None
            price_cols: list[int] = []
            total_cols: list[int] = []
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                val = normalize(_cell_text(ws.cell(row_idx, col_idx).value))
                if "вид работы" in val:
                    work_col = col_idx
                elif "материал" in val:
                    material_col = col_idx
                if "цен" in val or "стоим" in val:
                    price_cols.append(col_idx)
                if "сумм" in val or "итого" in val:
                    total_cols.append(col_idx)
            if work_col and material_col:
                price_anchor = max(price_cols) if price_cols else material_col
                return {
                    "header_row": row_idx,
                    "work_col": work_col,
                    "material_col": material_col,
                    "rightmost_total_cols": [c for c in total_cols if c > price_anchor] or total_cols,
                }
        return None

    def _parse_type1(self, ws: Worksheet, info: dict) -> list[ParsedRow]:
        results: list[ParsedRow] = []
        current_section: Optional[str] = None
        current_work: ParsedRow | None = None
        pending_materials: list[dict] = []
        order = 0

        for row_idx in range(info["header_row"] + 1, ws.max_row + 1):
            num = ws.cell(row_idx, info["num"]).value
            name = _cell_text(ws.cell(row_idx, info["name"]).value)
            tipo = _cell_text(ws.cell(row_idx, info["type"]).value).lower()
            unit = ws.cell(row_idx, info.get("unit", 0)).value if info.get("unit") else None
            qty = ws.cell(row_idx, info.get("qty", 0)).value if info.get("qty") else None
            price = ws.cell(row_idx, info.get("price", 0)).value if info.get("price") else None
            total = ws.cell(row_idx, info.get("total", 0)).value if info.get("total") else None

            if not any((num, name, tipo, unit, qty, price, total)):
                continue

            num_str = _cell_text(num)
            if self._is_group_row_type1(ws, row_idx, info["name"], num_str, name, unit, qty):
                current_section = name or current_section
                current_work = None
                pending_materials.clear()
                continue

            row_kind = self._detect_type1_row_kind(num_str, tipo, unit, qty, total)
            if row_kind == "material" and name:
                material = _make_material(name, unit=unit, qty=qty, price=price, total=total)
                if current_work is not None:
                    current_work.materials.append(material)
                else:
                    pending_materials.append(material)
                continue

            if row_kind != "work" or not name:
                continue

            current_work = ParsedRow(
                section=current_section,
                work_name=name,
                unit=_to_str(unit),
                quantity=_to_optional_float(qty),
                unit_price=_to_optional_float(price, treat_zero_as_none=True),
                total_price=_to_optional_float(total, treat_zero_as_none=True),
                materials=list(pending_materials),
                row_order=order,
                raw_data={
                    "num": num_str,
                    "type": tipo or "work",
                },
                source_strategy=self.name,
            )
            pending_materials.clear()
            results.append(current_work)
            order += 1

        return results

    def _parse_type2(self, ws: Worksheet, info: dict) -> list[ParsedRow]:
        results: list[ParsedRow] = []
        current_section: Optional[str] = None
        current_work: ParsedRow | None = None
        pending_materials: list[dict] = []
        order = 0

        base = info["work_col"]
        cols = {
            "work_name": base,
            "material_name": base + 1,
            "work_unit": base + 2,
            "material_unit": base + 3,
            "work_qty": base + 4,
            "material_qty": base + 5,
            "work_price": base + 6,
            "work_price_markup": base + 7,
            "material_price": base + 8,
            "work_total": base + 9,
            "material_total": base + 10,
            "grand_total": base + 11,
        }

        for row_idx in range(info["header_row"] + 1, ws.max_row + 1):
            work_name = _cell_text(ws.cell(row_idx, cols["work_name"]).value)
            material_name = _cell_text(ws.cell(row_idx, cols["material_name"]).value)
            work_unit = ws.cell(row_idx, cols["work_unit"]).value
            material_unit = ws.cell(row_idx, cols["material_unit"]).value
            work_qty = ws.cell(row_idx, cols["work_qty"]).value
            material_qty = ws.cell(row_idx, cols["material_qty"]).value
            work_price = ws.cell(row_idx, cols["work_price"]).value
            work_price_markup = ws.cell(row_idx, cols["work_price_markup"]).value
            material_price = ws.cell(row_idx, cols["material_price"]).value
            work_total = ws.cell(row_idx, cols["work_total"]).value
            material_total = ws.cell(row_idx, cols["material_total"]).value
            grand_total = ws.cell(row_idx, cols["grand_total"]).value
            row_rightmost_total = _pick_rightmost_numeric(ws, row_idx, info.get("rightmost_total_cols", []))

            if not work_name and not material_name:
                continue

            if self._is_group_row_type2(ws, row_idx, cols["work_name"], work_name, material_name, work_unit, work_qty, work_total):
                current_section = work_name or current_section
                current_work = None
                pending_materials.clear()
                continue

            if work_name and self._is_work_row_type2(work_name, work_unit, work_qty, work_total, work_price, material_name):
                current_work = ParsedRow(
                    section=current_section,
                    work_name=work_name,
                    unit=_to_str(work_unit),
                    quantity=_to_optional_float(work_qty),
                    unit_price=_to_optional_float(work_price, treat_zero_as_none=True)
                    or _to_optional_float(work_price_markup, treat_zero_as_none=True),
                    total_price=_to_optional_float(row_rightmost_total, treat_zero_as_none=True)
                    or _to_optional_float(work_total, treat_zero_as_none=True)
                    or _to_optional_float(grand_total, treat_zero_as_none=True),
                    materials=list(pending_materials),
                    row_order=order,
                    raw_data={},
                    source_strategy=self.name,
                )
                pending_materials.clear()
                results.append(current_work)
                order += 1
                continue

            if material_name:
                material = _make_material(
                    material_name,
                    unit=material_unit,
                    qty=material_qty,
                    price=material_price,
                    total=material_total or grand_total,
                )
                if current_work is not None:
                    current_work.materials.append(material)
                else:
                    pending_materials.append(material)

        return results

    def _is_group_row_type1(
        self,
        ws: Worksheet,
        row_idx: int,
        name_col: int,
        num: str,
        name: str,
        unit,
        qty,
    ) -> bool:
        if not name:
            return False
        if num.isdigit():
            return True
        return _is_empty_metric(unit) and _is_empty_metric(qty) and (
            _cell_is_emphasized(ws, row_idx, name_col) or _looks_like_group_label(name)
        )

    def _detect_type1_row_kind(self, num: str, tipo: str, unit, qty, total) -> str:
        if "матер" in tipo or tipo.strip() == "материал":
            return "material"
        if _TYPE1_ITEM_RE.match(num):
            return "work"
        return "work" if (unit or not _is_empty_metric(qty) or not _is_empty_metric(total) or tipo) else "other"

    def _is_group_row_type2(
        self,
        ws: Worksheet,
        row_idx: int,
        work_col: int,
        work_name: str,
        material_name: str,
        work_unit,
        work_qty,
        work_total,
    ) -> bool:
        if not work_name or material_name:
            return False
        if not _is_empty_metric(work_qty) or not _is_empty_metric(work_total) or _to_str(work_unit):
            return False
        return _cell_is_emphasized(ws, row_idx, work_col) or _looks_like_group_label(work_name)

    def _is_work_row_type2(
        self,
        work_name: str,
        work_unit,
        work_qty,
        work_total,
        work_price,
        material_name: str,
    ) -> bool:
        if not work_name:
            return False
        if not material_name and not any(
            (
                _to_str(work_unit),
                not _is_empty_metric(work_qty),
                not _is_empty_metric(work_total),
                not _is_empty_metric(work_price),
            )
        ):
            return False
        return True


# ─────────────────────────────────────────────────────────────────────────────
# STRATEGY PROTOCOL
# ─────────────────────────────────────────────────────────────────────────────

class ParserStrategy(Protocol):
    def can_parse(self, ws: Worksheet) -> float:
        """Возвращает уверенность [0.0 – 1.0], что этот формат подходит."""
        ...

    def parse(self, ws: Worksheet) -> list[ParsedRow]:
        ...


# ─────────────────────────────────────────────────────────────────────────────
# СТРАТЕГИЯ 1: СТРОЧНАЯ (классическая таблица)
# ─────────────────────────────────────────────────────────────────────────────

class RowOrientedParser:
    """
    Формат:
        Строка 1: ... (возможно пустая / шапка)
        Строка N: [Наименование] [Ед.] [Кол-во] [Цена] [Сумма]   ← заголовки
        Строка N+1...: данные
    """

    name = "row"

    def can_parse(self, ws: Worksheet) -> float:
        header_row = self._find_header_row(ws)
        if header_row is None:
            return 0.0
        col_map = self._map_columns(ws, header_row)
        # Уверенность растёт с количеством найденных колонок
        found = len([v for v in col_map.values() if v is not None])
        return min(1.0, found / 3)   # нашли ≥3 колонок → уверенность 1.0

    def parse(self, ws: Worksheet) -> list[ParsedRow]:
        header_row = self._find_header_row(ws)
        if header_row is None:
            return []

        col_map = self._map_columns(ws, header_row)
        rows = self._extract_rows(ws, header_row, col_map)
        return rows

    # ── internals ─────────────────────────────────────────────────────────────

    def _find_header_row(self, ws: Worksheet) -> Optional[int]:
        """
        Ищем строку, в которой ≥2 ячеек совпадают с нашими алиасами.
        Сканируем первые 40 строк.
        """
        for row_idx in range(1, min(41, ws.max_row + 1)):
            hits = 0
            for col_idx in range(1, min(ws.max_column + 1, 30)):
                cell = ws.cell(row_idx, col_idx).value
                if cell and match_any_field(str(cell)):
                    hits += 1
            if hits >= 2:
                return row_idx
        return None

    def _map_columns(self, ws: Worksheet, header_row: int) -> dict[str, Optional[int]]:
        """
        Возвращает {field_name: col_index (1-based)} для всех полей.
        Если колонка не найдена — None.
        """
        col_map: dict[str, Optional[int]] = {f: None for f in ALIASES}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(header_row, col_idx).value
            if not cell_val:
                continue
            field = match_any_field(str(cell_val))
            if field and col_map[field] is None:  # первое совпадение побеждает
                col_map[field] = col_idx
        return col_map

    def _extract_rows(
        self, ws: Worksheet, header_row: int, col_map: dict
    ) -> list[ParsedRow]:
        results: list[ParsedRow] = []
        current_section: Optional[str] = None
        order = 0

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_values = {
                field: ws.cell(row_idx, col).value
                for field, col in col_map.items()
                if col is not None
            }

            # Пропускаем полностью пустые строки
            if not any(v for v in row_values.values() if v is not None):
                continue

            work_name = row_values.get("work_name")

            # Определяем: это раздел или строка данных?
            if self._is_section(row_values, col_map):
                current_section = str(work_name).strip() if work_name else current_section
                continue

            if not work_name:
                continue

            results.append(ParsedRow(
                section=current_section,
                work_name=str(work_name).strip(),
                unit=_to_str(row_values.get("unit")),
                quantity=_to_float(row_values.get("quantity")),
                unit_price=_to_float(row_values.get("unit_price")),
                total_price=_to_float(row_values.get("total_price")),
                row_order=order,
                raw_data={f: str(v) for f, v in row_values.items() if v is not None},
                source_strategy="row",
            ))
            order += 1

        return results

    def _is_section(self, row_values: dict, col_map: dict) -> bool:
        """
        Раздел = есть work_name, но нет числовых значений.
        Дополнительно: текст не слишком длинный (< 120 символов).
        """
        work_name = row_values.get("work_name")
        if not work_name:
            return False
        has_numbers = any(
            _to_float(row_values.get(f)) is not None
            for f in ("quantity", "unit_price", "total_price")
        )
        return not has_numbers and len(str(work_name)) < 120


# ─────────────────────────────────────────────────────────────────────────────
# СТРАТЕГИЯ 2: СТОЛБЦОВАЯ
# ─────────────────────────────────────────────────────────────────────────────

class ColumnOrientedParser:
    """
    Формат (встречается в Excel-сметах подрядчиков):

        Строка 1: Наименование  | Монтаж кровли | Устройство стяжки | ...
        Строка 2: Ед.изм        | м2            | м2                | ...
        Строка 3: Количество    | 450           | 320               | ...
        Строка 4: Цена за ед.   | 850           | 400               | ...
        Строка 5: Сумма         | 382500        | 128000            | ...

    Т.е. первая колонка = метки строк, остальные = отдельные работы.
    """

    name = "column"
    LABEL_COL = 1       # колонка с метками (1-based)
    DATA_START_COL = 2  # с какой колонки начинаются данные

    def can_parse(self, ws: Worksheet) -> float:
        label_fields = self._detect_label_rows(ws)
        # Если в первой колонке нашли ≥3 алиаса → скорее всего столбцовый формат
        return min(1.0, len(label_fields) / 3)

    def parse(self, ws: Worksheet) -> list[ParsedRow]:
        label_map = self._detect_label_rows(ws)
        if not label_map:
            return []

        results: list[ParsedRow] = []
        order = 0

        for col_idx in range(self.DATA_START_COL, ws.max_column + 1):
            work_name_row = label_map.get("work_name")
            work_name_val = ws.cell(work_name_row, col_idx).value if work_name_row else None

            if not work_name_val:
                continue

            def get(field: str):
                row = label_map.get(field)
                return ws.cell(row, col_idx).value if row else None

            results.append(ParsedRow(
                section=None,          # в столбцовом формате разделы обычно отсутствуют
                work_name=str(work_name_val).strip(),
                unit=_to_str(get("unit")),
                quantity=_to_float(get("quantity")),
                unit_price=_to_float(get("unit_price")),
                total_price=_to_float(get("total_price")),
                row_order=order,
                raw_data={
                    field: str(ws.cell(row, col_idx).value)
                    for field, row in label_map.items()
                    if ws.cell(row, col_idx).value is not None
                },
                source_strategy="column",
            ))
            order += 1

        return results

    # ── internals ─────────────────────────────────────────────────────────────

    def _detect_label_rows(self, ws: Worksheet) -> dict[str, int]:
        """
        Сканируем первую колонку — ищем строки, значения которых
        соответствуют нашим алиасам.
        Возвращает {field_name: row_index}.
        """
        label_map: dict[str, int] = {}
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            cell_val = ws.cell(row_idx, self.LABEL_COL).value
            if not cell_val:
                continue
            field = match_any_field(str(cell_val))
            if field and field not in label_map:
                label_map[field] = row_idx
        return label_map


# ─────────────────────────────────────────────────────────────────────────────
# СТРАТЕГИЯ 3: СМЕШАННАЯ (блочная)
# ─────────────────────────────────────────────────────────────────────────────

class BlockOrientedParser:
    """
    Формат: несколько независимых таблиц на одном листе.
    Каждый «блок» — это мини-таблица со своей шапкой.
    Встречается в сметах, скомпонованных вручную.

    Алгоритм:
      - Ищем все строки, которые могут быть заголовками (≥2 хитов)
      - Для каждого найденного заголовка парсим блок до следующего
    """

    name = "block"

    def can_parse(self, ws: Worksheet) -> float:
        header_rows = self._find_all_header_rows(ws)
        return 0.8 if len(header_rows) > 1 else 0.0

    def parse(self, ws: Worksheet) -> list[ParsedRow]:
        header_rows = self._find_all_header_rows(ws)
        if not header_rows:
            return []

        all_results: list[ParsedRow] = []
        current_section: Optional[str] = None
        row_parser = RowOrientedParser()

        for i, header_row in enumerate(header_rows):
            # Граница блока: до следующего заголовка или конца листа
            end_row = header_rows[i + 1] - 1 if i + 1 < len(header_rows) else ws.max_row

            # Определяем название раздела: ищем непустую строку выше заголовка
            for r in range(header_row - 1, max(0, header_row - 5), -1):
                candidate = ws.cell(r, 1).value
                if candidate and not match_any_field(str(candidate)):
                    current_section = str(candidate).strip()
                    break

            # Парсим блок как обычную строчную таблицу
            col_map = row_parser._map_columns(ws, header_row)
            block_rows = self._extract_block(ws, header_row, end_row, col_map, current_section)
            all_results.extend(block_rows)

        return all_results

    def _find_all_header_rows(self, ws: Worksheet) -> list[int]:
        found = []
        for row_idx in range(1, ws.max_row + 1):
            hits = sum(
                1 for col_idx in range(1, min(ws.max_column + 1, 30))
                if (cell := ws.cell(row_idx, col_idx).value) and match_any_field(str(cell))
            )
            if hits >= 2:
                found.append(row_idx)
        return found

    def _extract_block(
        self,
        ws: Worksheet,
        header_row: int,
        end_row: int,
        col_map: dict,
        section: Optional[str],
    ) -> list[ParsedRow]:
        results = []
        order = 0
        row_parser = RowOrientedParser()

        for row_idx in range(header_row + 1, end_row + 1):
            row_values = {
                field: ws.cell(row_idx, col).value
                for field, col in col_map.items()
                if col is not None
            }
            if not any(v for v in row_values.values() if v is not None):
                continue
            work_name = row_values.get("work_name")
            if not work_name:
                continue
            if row_parser._is_section(row_values, col_map):
                continue

            results.append(ParsedRow(
                section=section,
                work_name=str(work_name).strip(),
                unit=_to_str(row_values.get("unit")),
                quantity=_to_float(row_values.get("quantity")),
                unit_price=_to_float(row_values.get("unit_price")),
                total_price=_to_float(row_values.get("total_price")),
                row_order=order,
                raw_data={f: str(v) for f, v in row_values.items() if v is not None},
                source_strategy="block",
            ))
            order += 1

        return results


# ─────────────────────────────────────────────────────────────────────────────
# СТРАТЕГИЯ: ШАБЛОННЫЙ ЛИСТ (Смета / Estimate sheet)
# ─────────────────────────────────────────────────────────────────────────────

_SMETA_SHEET_RE = re.compile(r'смет|estimate|smeta', re.IGNORECASE)
_ITEM_CODE_RE   = re.compile(r'^\d+\.\d+$')      # «1.1», «2.10», «15.3»
_SECTION_HDR_RE = re.compile(r'^\d+\.\s+.+')     # «6. Работы по потолкам»
_SKIP_NAME_RE   = re.compile(
    r'^(ИТОГО|ОБЩАЯ СТОИМОСТЬ|СКИДКА|можно вписать|наименование работ)',
    re.IGNORECASE,
)

# Фиксированные колонки шаблона (0-based): A=0 пустая, B=1 код, C=2 название, D=3 ед.изм.
# Колонки qty/price/total определяются динамически из строки заголовка.
_TC_CODE = 1
_TC_NAME = 2
_TC_UNIT = 3

# Алиасы заголовков для динамического поиска qty / price / total
_HDR_QTY   = re.compile(r'^кол', re.IGNORECASE)
_HDR_PRICE = re.compile(r'^цена$', re.IGNORECASE)
_HDR_TOTAL = re.compile(r'^сумма$', re.IGNORECASE)


def _detect_template_cols(ws, max_scan: int = 10) -> tuple:
    """
    Ищет строку-заголовок в первых max_scan строках листа.
    Возвращает (idx_qty, idx_price, idx_total) — 0-based индексы в строке.
    Если заголовок не найден — возвращает дефолтные значения (4, 5, 6).
    """
    default = (4, 5, 6)   # E, F, G — исходный шаблон без «Базовой Цены»

    for raw_row in ws.iter_rows(min_col=1, max_col=10, max_row=max_scan, values_only=True):
        row = [str(v).strip() if v is not None else "" for v in raw_row]
        qty_idx = price_idx = total_idx = None
        for i, cell in enumerate(row):
            if _HDR_QTY.match(cell):
                qty_idx = i
            elif _HDR_PRICE.match(cell):
                price_idx = i
            elif _HDR_TOTAL.match(cell):
                total_idx = i
        if qty_idx is not None and total_idx is not None:
            if price_idx is None:
                price_idx = total_idx - 1
            return (qty_idx, price_idx, total_idx)

    return default


class TemplateSheetStrategy:
    """
    Парсит Excel-сметы на основе именованного листа (Смета / Estimate).

    Признаки формата:
    - Есть лист с именем, содержащим «смет» / «estimate»
    - Фиксированные колонки: B=код, C=название, D=ед.изм.
    - Колонки qty / price / total определяются из строки заголовка —
      корректно разбирает шаблоны с дополнительными колонками
      (например «Базовая Цена» между ед.изм. и кол-вом)
    - max_column может быть огромным из-за merged cells — читаем только
      нужное количество колонок (read_only=True)
    - Позиции с total=0 и qty=0/None — незаполненные строки прайс-листа,
      пропускаем
    """

    name = "smeta_template"

    def can_parse(self, ws) -> float:
        return 0.95 if _SMETA_SHEET_RE.search(ws.title or "") else 0.0

    def parse(self, ws) -> list[ParsedRow]:
        result: list[ParsedRow] = []
        section: Optional[str] = None
        order = 0

        idx_qty, idx_price, idx_total = _detect_template_cols(ws)
        max_col = max(idx_qty, idx_price, idx_total) + 2

        for raw_row in ws.iter_rows(min_col=1, max_col=max_col, values_only=True):
            row = list(raw_row) + [None] * (max_col - len(raw_row))

            code  = row[_TC_CODE]
            name  = row[_TC_NAME]
            unit  = row[_TC_UNIT]
            qty   = _to_float(row[idx_qty])
            price = _to_float(row[idx_price])
            total = _to_float(row[idx_total])

            name_str = str(name).strip() if name else ""
            if not name_str:
                continue
            if _SKIP_NAME_RE.match(name_str):
                continue

            code_str = str(code).strip() if code else ""

            # Заголовок раздела: нет кода, имя вида «N. Название»
            if not code_str and _SECTION_HDR_RE.match(name_str):
                section = re.sub(r'^\d+\.\s*', '', name_str).strip()
                continue

            # Строка данных — обязателен числовой код
            if not _ITEM_CODE_RE.match(code_str):
                continue

            # Незаполненная позиция прайс-листа — пропускаем
            if (total is None or total == 0) and (qty is None or qty == 0):
                continue

            result.append(ParsedRow(
                section         = section,
                work_name       = name_str,
                unit            = _to_str(unit),
                quantity        = qty,
                unit_price      = price,
                total_price     = total if total else None,
                row_order       = order,
                raw_data        = {"code": code_str},
                source_strategy = "smeta_template",
            ))
            order += 1

        return result


# ─────────────────────────────────────────────────────────────────────────────
# ПАРСИНГ С РУЧНЫМ МАППИНГОМ КОЛОНОК
# ─────────────────────────────────────────────────────────────────────────────

# Ключи маппинга, которые ожидаем от фронта
MAPPING_FIELDS = ("work_name", "unit", "quantity", "unit_price", "total_price")


def parse_with_mapping(
    file_path: str | Path,
    col_mapping: dict[int, str],   # {col_index_0based: field_key | "skip"}
    sheet: Optional[str] = None,
    preview_rows: int = 3,
) -> list[ParsedRow]:
    """
    Парсит Excel по заданному маппингу колонок.
    col_mapping пример: {0: "work_name", 1: "unit", 2: "quantity", 3: "unit_price", 4: "total_price"}
    Колонки с "skip" игнорируются.
    """
    file_path = Path(file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    results: list[ParsedRow] = []
    order = 0

    # Определяем, с какой строки начинаются данные:
    # пропускаем строки где в «work_name» колонке стоит что-то похожее на заголовок
    work_col = next((c for c, f in col_mapping.items() if f == "work_name"), None)

    start_row = 1
    if work_col is not None:
        for row_idx in range(1, min(10, ws.max_row + 1)):
            val = ws.cell(row_idx, work_col + 1).value  # openpyxl 1-based
            if val and match_any_field(str(val)):
                start_row = row_idx + 1
                break

    for row_idx in range(start_row, ws.max_row + 1):
        row_data: dict[str, any] = {}
        for col_0, field_key in col_mapping.items():
            if field_key == "skip":
                continue
            row_data[field_key] = ws.cell(row_idx, col_0 + 1).value  # openpyxl 1-based

        # Пропускаем пустые строки
        if not any(v for v in row_data.values() if v is not None):
            continue

        work_name = row_data.get("work_name")
        if not work_name:
            continue

        results.append(ParsedRow(
            work_name   = str(work_name).strip(),
            unit        = _to_str(row_data.get("unit")),
            quantity    = _to_float(row_data.get("quantity")),
            unit_price  = _to_float(row_data.get("unit_price")),
            total_price = _to_float(row_data.get("total_price")),
            row_order   = order,
            raw_data    = {k: str(v) for k, v in row_data.items() if v is not None},
            source_strategy = "manual_mapping",
        ))
        order += 1

    return results


# ─────────────────────────────────────────────────────────────────────────────
# ГЛАВНЫЙ ПАРСЕР — выбирает стратегию автоматически
# ─────────────────────────────────────────────────────────────────────────────

# Порог уверенности: ниже него → просим пользователя сделать маппинг вручную
CONFIDENCE_THRESHOLD = 0.8
# Количество строк превью для UI маппинга
PREVIEW_ROWS_COUNT = 3


class ExcelEstimateParser:
    """
    Использование (авто):
        parser = ExcelEstimateParser()
        rows, meta = parser.parse("smeta.xlsx")
        # Может поднять NeedsMappingError — тогда показываем UI маппинга

    Использование (после ручного маппинга):
        rows = parser.parse_mapped("smeta.xlsx", col_mapping={0: "work_name", ...})

    meta содержит:
        {
            "strategy":   "row" | "column" | "block" | "smeta_template",
            "sheet":      "Sheet1",
            "confidence": 0.95,
            "rows_found": 47,
        }
    """

    STRATEGIES: list = [
        RowOrientedParser(),
        ColumnOrientedParser(),
        BlockOrientedParser(),
    ]

    def parse(self, file_path: str | Path) -> tuple[list[ParsedRow], dict]:
        file_path = Path(file_path)

        # ── Приоритетный путь: сначала определяем явный тип сметы и разбираем
        # его отдельным алгоритмом до общих эвристик. Это нужно для файлов,
        # где есть явные указатели "работы/материалы".
        wb_structured = openpyxl.load_workbook(file_path, data_only=True)
        structured = StructuredSmetaParser()
        for sheet_name in wb_structured.sheetnames:
            ws_structured = wb_structured[sheet_name]
            confidence = structured.can_parse(ws_structured)
            if confidence >= CONFIDENCE_THRESHOLD:
                rows = structured.parse(ws_structured)
                if rows:
                    wb_structured.close()
                    return rows, {
                        "strategy": structured.name,
                        "sheet": sheet_name,
                        "confidence": confidence,
                        "rows_found": len(rows),
                    }
        wb_structured.close()

        # ── Быстрая проверка: есть ли лист типа «Смета»? ──────────────────
        wb_ro = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        smeta_sheet = next(
            (s for s in wb_ro.sheetnames if _SMETA_SHEET_RE.search(s)), None
        )
        wb_ro.close()

        if smeta_sheet:
            wb_named = openpyxl.load_workbook(str(file_path), data_only=True)
            ws_named = wb_named[smeta_sheet]

            structured = StructuredSmetaParser()
            if structured.can_parse(ws_named) >= CONFIDENCE_THRESHOLD:
                rows = structured.parse(ws_named)
                wb_named.close()
                if rows:
                    return rows, {
                        "strategy":   structured.name,
                        "sheet":      smeta_sheet,
                        "confidence": 0.99,
                        "rows_found": len(rows),
                    }

            wb_named.close()

            wb_ro = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            ws = wb_ro[smeta_sheet]
            strategy = TemplateSheetStrategy()
            rows = strategy.parse(ws)
            wb_ro.close()
            if rows:
                return rows, {
                    "strategy":   strategy.name,
                    "sheet":      smeta_sheet,
                    "confidence": 0.95,
                    "rows_found": len(rows),
                }

        # ── Стандартный путь для остальных форматов ───────────────────────
        wb = openpyxl.load_workbook(file_path, data_only=True)
        best_rows: list[ParsedRow] = []
        best_meta = {"strategy": "none", "sheet": "", "confidence": 0.0}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 3 or ws.max_column < 2:
                continue

            for strategy in self.STRATEGIES:
                confidence = strategy.can_parse(ws)
                if confidence > best_meta["confidence"]:
                    try:
                        rows = strategy.parse(ws)
                        if rows:
                            best_rows = rows
                            best_meta = {
                                "strategy":   strategy.name,
                                "sheet":      sheet_name,
                                "confidence": confidence,
                                "rows_found": len(rows),
                            }
                    except Exception:
                        continue

        # ── Уверенность низкая — просим ручной маппинг ────────────────────
        if best_meta["confidence"] < CONFIDENCE_THRESHOLD:
            sheet_name = best_meta.get("sheet") or wb.sheetnames[0]
            ws = wb[sheet_name]
            preview = self._get_preview_rows(ws, n=PREVIEW_ROWS_COUNT)
            raise NeedsMappingError(
                filename     = file_path.name,
                sheet        = sheet_name,
                preview_rows = preview,
                col_count    = len(preview[0]) if preview else 0,
            )

        if not best_rows:
            raise ValueError(
                "Не удалось распознать формат сметы. "
                "Убедитесь, что файл содержит колонки: "
                "наименование, количество, единица измерения, сумма."
            )

        return best_rows, best_meta

    def parse_mapped(
        self,
        file_path: str | Path,
        col_mapping: dict[int, str],
        sheet: Optional[str] = None,
    ) -> tuple[list[ParsedRow], dict]:
        """Парсит файл с явным маппингом колонок (после ручного назначения)."""
        rows = parse_with_mapping(file_path, col_mapping, sheet=sheet)
        if not rows:
            raise ValueError("После применения маппинга не найдено ни одной строки.")
        return rows, {
            "strategy":   "manual_mapping",
            "sheet":      sheet or "auto",
            "confidence": 1.0,
            "rows_found": len(rows),
        }

    def _get_preview_rows(self, ws: Worksheet, n: int = 3) -> list[list]:
        """
        Возвращает первые n непустых строк данных из листа.
        Нормализует значения к строкам для удобной сериализации.
        """
        results = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(v).strip() if v is not None else "" for v in row]
            # Пропускаем полностью пустые строки
            if not any(c for c in cells):
                continue
            # Пропускаем строки-заголовки (где много совпадений с алиасами)
            hits = sum(1 for c in cells if c and match_any_field(c))
            if hits >= 2:
                continue
            results.append(cells)
            if len(results) >= n:
                break
        return results


# ─────────────────────────────────────────────────────────────────────────────
# УТИЛИТЫ
# ─────────────────────────────────────────────────────────────────────────────

def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d,.\-]", "", str(value)).replace(",", ".")
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None
