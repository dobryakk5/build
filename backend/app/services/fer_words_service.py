from __future__ import annotations

import asyncio
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable, Sequence
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Estimate, EstimateBatch, FerWordsEntry, Job

try:
    from rapidfuzz import fuzz as rapidfuzz_fuzz
except ImportError:  # pragma: no cover - fallback for local env without rapidfuzz
    rapidfuzz_fuzz = None


FER_WORDS_TEXT_COLUMNS = tuple(range(3, 12))  # C:K
FER_WORDS_CODE_COLUMN = 2  # B
FER_WORDS_HUMAN_HOURS_COLUMN = 13  # M
FER_WORDS_MACHINE_HOURS_COLUMN = 14  # N

LOW_SCORE_THRESHOLD = 0.88
LOW_MATCH_COUNT = 2
STOPWORDS = {
    "в",
    "во",
    "и",
    "или",
    "для",
    "до",
    "от",
    "по",
    "при",
    "с",
    "со",
    "на",
    "из",
    "над",
    "под",
    "к",
    "ко",
    "а",
}


@dataclass(slots=True)
class FerWordsCandidate:
    entry_id: int
    fer_code: str
    display_name: str
    human_hours: float | None
    machine_hours: float | None
    matched_tokens: int
    exact_matches: int
    numeric_matches: int
    average_ratio: float
    score: float

    def to_payload(self) -> dict[str, object]:
        return {
            "entry_id": self.entry_id,
            "fer_code": self.fer_code,
            "display_name": self.display_name,
            "human_hours": self.human_hours,
            "machine_hours": self.machine_hours,
            "matched_tokens": self.matched_tokens,
            "exact_matches": self.exact_matches,
            "numeric_matches": self.numeric_matches,
            "average_ratio": round(self.average_ratio, 4),
            "score": round(self.score, 4),
        }


def _ratio(left: str, right: str) -> int:
    if rapidfuzz_fuzz is not None:
        return int(rapidfuzz_fuzz.ratio(left, right))
    return int(SequenceMatcher(a=left, b=right).ratio() * 100)


def normalize_fer_words_text(value: str | None) -> str:
    if not value:
        return ""

    text = str(value).lower().replace("ё", "е")
    text = re.sub(r"(\d),(\d)", r"\1.\2", text)
    text = text.replace("м3", " м3 ").replace("м2", " м2 ").replace("мп", " мп ")
    text = re.sub(r"(\d)([a-zа-я])", r"\1 \2", text, flags=re.IGNORECASE)
    text = re.sub(r"([a-zа-я])(\d)", lambda match: match.group(0) if match.group(0) in {"м2", "м3", "мп"} else f"{match.group(1)} {match.group(2)}", text, flags=re.IGNORECASE)
    text = re.sub(r"[()\"'«»:/;,+]", " ", text)
    text = re.sub(r"[^\w.\-]+", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokenize_fer_words_text(value: str | None) -> list[str]:
    normalized = normalize_fer_words_text(value)
    tokens: list[str] = []
    for token in normalized.split():
        cleaned = token.strip("._-")
        if not cleaned:
            continue
        if cleaned in STOPWORDS:
            continue
        if len(cleaned) == 1 and not cleaned.isdigit():
            continue
        tokens.append(cleaned)
    return tokens


def build_fer_words_display_name(parts: Sequence[str | None]) -> str:
    return " · ".join(str(part).strip() for part in parts if part and str(part).strip())


def build_estimate_fer_words_text(estimate: Estimate) -> str:
    parts = [str(estimate.work_name).strip()]
    if estimate.section and str(estimate.section).strip():
        parts.insert(0, str(estimate.section).strip())
    if estimate.unit and str(estimate.unit).strip():
        parts.append(str(estimate.unit).strip())
    return " ".join(parts)


def _is_numeric_token(token: str) -> bool:
    return any(ch.isdigit() for ch in token)


def score_fer_words_candidate(
    estimate_tokens: Sequence[str],
    candidate_tokens: Sequence[str],
) -> tuple[int, int, int, float, float]:
    if not estimate_tokens or not candidate_tokens:
        return 0, 0, 0, 0.0, 0.0

    matched_tokens = 0
    exact_matches = 0
    numeric_matches = 0
    ratio_sum = 0.0
    used_indexes: set[int] = set()

    for estimate_token in estimate_tokens:
        best_ratio = 0
        best_index: int | None = None

        for index, candidate_token in enumerate(candidate_tokens):
            if index in used_indexes:
                continue
            ratio = _ratio(estimate_token, candidate_token)
            if ratio > best_ratio:
                best_ratio = ratio
                best_index = index
            if ratio == 100:
                break

        if best_index is None or best_ratio < 84:
            continue

        used_indexes.add(best_index)
        matched_tokens += 1
        ratio_sum += best_ratio
        if best_ratio == 100:
            exact_matches += 1
        if _is_numeric_token(estimate_token) and estimate_token == candidate_tokens[best_index]:
            numeric_matches += 1

    if matched_tokens == 0:
        return 0, 0, 0, 0.0, 0.0

    average_ratio = ratio_sum / matched_tokens / 100.0
    score = matched_tokens + exact_matches * 0.01 + numeric_matches * 0.005 + average_ratio * 0.001
    return matched_tokens, exact_matches, numeric_matches, average_ratio, score


def build_fer_words_candidates(
    estimate_text: str,
    entries: Iterable[FerWordsEntry],
    limit: int = 5,
) -> list[FerWordsCandidate]:
    estimate_tokens = tokenize_fer_words_text(estimate_text)
    candidates: list[FerWordsCandidate] = []

    for entry in entries:
        candidate_tokens = list(entry.search_tokens or [])
        matched_tokens, exact_matches, numeric_matches, average_ratio, score = score_fer_words_candidate(
            estimate_tokens,
            candidate_tokens,
        )
        if matched_tokens == 0:
            continue
        candidates.append(
            FerWordsCandidate(
                entry_id=entry.id,
                fer_code=entry.fer_code,
                display_name=entry.display_name,
                human_hours=float(entry.human_hours) if entry.human_hours is not None else None,
                machine_hours=float(entry.machine_hours) if entry.machine_hours is not None else None,
                matched_tokens=matched_tokens,
                exact_matches=exact_matches,
                numeric_matches=numeric_matches,
                average_ratio=average_ratio,
                score=score,
            )
        )

    candidates.sort(
        key=lambda item: (
            item.matched_tokens,
            item.exact_matches,
            item.numeric_matches,
            item.average_ratio,
            item.score,
            item.fer_code,
        ),
        reverse=True,
    )
    return candidates[:limit]


def build_fer_words_candidate_for_entry(
    estimate_text: str,
    entry: FerWordsEntry,
) -> FerWordsCandidate | None:
    candidates = build_fer_words_candidates(estimate_text, [entry], limit=1)
    return candidates[0] if candidates else None


def should_auto_apply_fer_words(candidates: Sequence[FerWordsCandidate]) -> bool:
    if not candidates:
        return False
    top = candidates[0]
    if top.matched_tokens < LOW_MATCH_COUNT or top.average_ratio < LOW_SCORE_THRESHOLD:
        return False
    if len(candidates) > 1 and candidates[1].matched_tokens == top.matched_tokens:
        return False
    return True


async def import_fer_words_xlsx(db: AsyncSession, file_path: str) -> int:
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    entries: list[FerWordsEntry] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
            code = row[FER_WORDS_CODE_COLUMN - 1]
            if code is None or not str(code).strip():
                continue

            parts = [
                str(row[index - 1]).strip() if row[index - 1] is not None and str(row[index - 1]).strip() else None
                for index in FER_WORDS_TEXT_COLUMNS
            ]
            display_name = build_fer_words_display_name(parts)
            if not display_name:
                continue

            normalized_text = normalize_fer_words_text(display_name)
            search_tokens = tokenize_fer_words_text(display_name)
            entries.append(
                FerWordsEntry(
                    source_filename=path.name,
                    source_sheet=sheet_name,
                    source_row_number=row_number,
                    fer_code=str(code).strip(),
                    display_name=display_name,
                    normalized_text=normalized_text,
                    search_tokens=search_tokens,
                    human_hours=_coerce_number(row[FER_WORDS_HUMAN_HOURS_COLUMN - 1]),
                    machine_hours=_coerce_number(row[FER_WORDS_MACHINE_HOURS_COLUMN - 1]),
                    part_1=parts[0],
                    part_2=parts[1],
                    part_3=parts[2],
                    part_4=parts[3],
                    part_5=parts[4],
                    part_6=parts[5],
                    part_7=parts[6],
                    part_8=parts[7],
                    part_9=parts[8],
                )
            )

    await db.execute(FerWordsEntry.__table__.delete())  # type: ignore[arg-type]
    db.add_all(entries)
    await db.commit()
    return len(entries)


def _coerce_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


async def get_all_fer_words_entries(db: AsyncSession) -> list[FerWordsEntry]:
    return list(await db.scalars(select(FerWordsEntry).order_by(FerWordsEntry.id)))


async def get_fer_words_candidates_for_estimate(
    db: AsyncSession,
    estimate: Estimate,
    limit: int = 5,
) -> list[FerWordsCandidate]:
    entries = await get_all_fer_words_entries(db)
    estimate_text = build_estimate_fer_words_text(estimate)
    return build_fer_words_candidates(estimate_text, entries, limit=limit)


def apply_fer_words_choice(estimate: Estimate, entry: FerWordsEntry | None, candidate: FerWordsCandidate | None) -> None:
    if entry is None or candidate is None:
        estimate.fer_words_entry_id = None
        estimate.fer_words_code = None
        estimate.fer_words_name = None
        estimate.fer_words_human_hours = None
        estimate.fer_words_machine_hours = None
        estimate.fer_words_match_score = None
        estimate.fer_words_match_count = None
        estimate.fer_words_matched_at = None
        return

    estimate.fer_words_entry_id = entry.id
    estimate.fer_words_code = entry.fer_code
    estimate.fer_words_name = entry.display_name
    estimate.fer_words_human_hours = entry.human_hours
    estimate.fer_words_machine_hours = entry.machine_hours
    estimate.fer_words_match_score = round(candidate.average_ratio, 4)
    estimate.fer_words_match_count = candidate.matched_tokens
    estimate.fer_words_matched_at = datetime.now(timezone.utc)


async def start_fer_words_match_job(
    project_id: str,
    estimate_batch_id: str,
    user_id: str,
    db: AsyncSession,
) -> Job:
    batch = await db.scalar(
        select(EstimateBatch)
        .where(EstimateBatch.id == estimate_batch_id)
        .where(EstimateBatch.project_id == project_id)
        .where(EstimateBatch.deleted_at == None)
    )
    if batch is None:
        raise HTTPException(404, "Блок сметы не найден")

    estimate_count = await db.scalar(
        select(func.count())
        .select_from(Estimate)
        .where(Estimate.project_id == project_id)
        .where(Estimate.estimate_batch_id == estimate_batch_id)
        .where(Estimate.deleted_at == None)
    )
    if not estimate_count:
        raise HTTPException(400, "В выбранном блоке нет строк сметы")

    dictionary_count = await db.scalar(select(func.count()).select_from(FerWordsEntry))
    if not dictionary_count:
        raise HTTPException(400, "Справочник 'ФЕР слова' пуст. Сначала загрузите XLSX в базу.")

    job = Job(
        id=str(uuid4()),
        type="estimate_fer_words_match",
        status="pending",
        project_id=project_id,
        created_by=user_id,
        input={
            "estimate_batch_id": estimate_batch_id,
            "estimate_batch_name": batch.name,
        },
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)

    asyncio.create_task(_process_fer_words_match(job.id))
    return job


async def _process_fer_words_match(job_id: str) -> None:
    from app.core.database import AsyncSessionLocal

    async with AsyncSessionLocal() as db:
        job = await db.get(Job, job_id)
        if job is None:
            return

        estimate_batch_id = str(job.input.get("estimate_batch_id") or "")
        job.status = "processing"
        job.started_at = datetime.utcnow()
        await db.commit()

        try:
            entries = await get_all_fer_words_entries(db)
            estimates = list(
                await db.scalars(
                    select(Estimate)
                    .where(Estimate.project_id == job.project_id)
                    .where(Estimate.estimate_batch_id == estimate_batch_id)
                    .where(Estimate.deleted_at == None)
                    .order_by(Estimate.row_order)
                )
            )
            if not estimates:
                raise ValueError("В блоке сметы не найдено строк для сопоставления")

            matched_count = 0
            review_ids: list[str] = []

            for estimate in estimates:
                apply_fer_words_choice(estimate, None, None)

            entry_map = {entry.id: entry for entry in entries}
            for estimate in estimates:
                candidates = build_fer_words_candidates(
                    build_estimate_fer_words_text(estimate),
                    entries,
                    limit=5,
                )
                if not candidates:
                    continue
                top = candidates[0]
                if should_auto_apply_fer_words(candidates):
                    entry = entry_map.get(top.entry_id)
                    if entry is not None:
                        apply_fer_words_choice(estimate, entry, top)
                        matched_count += 1
                    continue
                review_ids.append(estimate.id)

            job.status = "done"
            job.result = {
                "estimate_batch_id": estimate_batch_id,
                "estimate_batch_name": job.input.get("estimate_batch_name"),
                "matched_rows_count": matched_count,
                "review_rows_count": len(review_ids),
                "review_estimate_ids": review_ids,
                "strategy": "fer_words",
            }
        except Exception as exc:
            job.status = "failed"
            job.result = {"error": str(exc), "strategy": "fer_words"}
        finally:
            job.finished_at = datetime.utcnow()
            await db.commit()
