"""Parser profile ``excel_sectioned_cost_split`` — КП/сметы с блоками
«РАБОТЫ … / МАТЕРИАЛЫ … / НАКЛАДНЫЕ …» и двухстрочной шапкой
«Стоимость работ | Стоимость материала | Всего» (напр. «грунтовые работы.xlsx»).

Тип строки задаётся блоком, в котором она лежит; ключевые слова уточняют его
(кран в блоке накладных — всё равно механизм, «Расходные материалы» — материал).
У строк-работ упомянутая техника дополнительно выносится отдельной строкой-механизмом.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from .excel_parser import _cell_text, _to_optional_float, _to_str, is_subtotal_label
from .pdf_parser import ParsedRow
from .resource_classifier import classify_sectioned_row, extract_mechanism_token

PROFILE_NAME = "excel_sectioned_cost_split"

_BLOCK_RE = re.compile(r"^\s*(РАБОТЫ|МАТЕРИАЛЫ|НАКЛАДНЫЕ)", re.IGNORECASE)
_BLOCK_TYPE = {"работы": "work", "материалы": "material", "накладные": "overhead"}

_NAME_HDR = re.compile(r"наименование.*(работ|материал)", re.IGNORECASE)
_WORK_COST_HDR = re.compile(r"стоимость\s+работ", re.IGNORECASE)
_MAT_COST_HDR = re.compile(r"стоимость\s+материал", re.IGNORECASE)


def detect(file_path: str | Path) -> bool:
    """True for the block + cost-split layout."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return False
    try:
        for ws in (wb[n] for n in wb.sheetnames):
            text_cells = []
            has_block = False
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True):
                for v in row:
                    if isinstance(v, str):
                        text_cells.append(v)
                        if _BLOCK_RE.match(v):
                            has_block = True
            blob = "\n".join(text_cells)
            if has_block and _WORK_COST_HDR.search(blob) and _MAT_COST_HDR.search(blob):
                return True
        return False
    finally:
        wb.close()


class ExcelSectionedCostSplitParser:
    name = PROFILE_NAME

    def parse(self, file_path: str | Path) -> tuple[list[ParsedRow], dict]:
        file_path = Path(file_path)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                cols = self._find_header(ws)
                if cols:
                    rows, declared = self._parse(ws, cols)
                    if rows:
                        return rows, {
                            "strategy": self.name,
                            "sheet": sheet_name,
                            "confidence": 0.95,
                            "rows_found": len(rows),
                            "declared_totals": declared,
                        }
        finally:
            wb.close()
        raise ValueError(
            "Профиль «Excel: блоки РАБОТЫ/МАТЕРИАЛЫ/НАКЛАДНЫЕ» не нашёл подходящую "
            "шапку (Наименование / Стоимость работ / Стоимость материала)."
        )

    def _find_header(self, ws) -> dict | None:
        """Locate the header columns by scanning the first rows (header spans 2-3
        rows: «Наименование…» + «Стоимость работ | Стоимость материала»)."""
        max_r = min(ws.max_row, 20)
        max_c = min(ws.max_column, 12)
        name_row = None
        cols: dict[str, int] = {}
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                t = _cell_text(ws.cell(r, c).value)
                if t and _NAME_HDR.search(t):
                    name_row = r
                    cols["name"] = c
                    break
            if name_row:
                break
        if not name_row:
            return None

        # Gather header text per column across the header band (name_row..+3).
        for c in range(1, max_c + 1):
            joined = " ".join(
                _cell_text(ws.cell(r, c).value)
                for r in range(name_row, min(name_row + 4, ws.max_row + 1))
            ).lower()
            if "ед" in joined and "изм" in joined:
                cols.setdefault("unit", c)
            elif "кол" in joined:
                cols.setdefault("qty", c)
            elif _WORK_COST_HDR.search(joined):
                cols["work_cost"] = c
            elif _MAT_COST_HDR.search(joined):
                cols["mat_cost"] = c
            elif "всего" in joined:
                cols.setdefault("total", c)

        if "work_cost" not in cols or "mat_cost" not in cols:
            return None
        cols["header_row"] = name_row
        return cols

    def _parse(self, ws, cols: dict) -> tuple[list[ParsedRow], list[dict]]:
        results: list[ParsedRow] = []
        declared: list[dict] = []
        block_default: str | None = None
        order = 0
        nm, uc, qc = cols["name"], cols.get("unit"), cols.get("qty")
        wcc, mcc, tc = cols["work_cost"], cols["mat_cost"], cols.get("total")

        for r in range(cols["header_row"] + 1, ws.max_row + 1):
            name = _cell_text(ws.cell(r, nm).value)
            work_cost = ws.cell(r, wcc).value
            mat_cost = ws.cell(r, mcc).value
            total = ws.cell(r, tc).value if tc else None
            unit = ws.cell(r, uc).value if uc else None
            qty = ws.cell(r, qc).value if qc else None

            # Subtotal rows («Итого …», «Всего по смете») — label may sit in the
            # work-cost column while the name column is empty.
            label = name or _cell_text(work_cost)
            if is_subtotal_label(label):
                value = _to_optional_float(mat_cost) or _to_optional_float(total)
                if value is not None:
                    declared.append({"section": None, "kind": "subtotal",
                                     "label": label, "total": value})
                continue

            if not name or name.strip().isdigit():
                continue

            has_metrics = any(_to_optional_float(v) is not None
                              for v in (qty, work_cost, mat_cost, total))

            # Block marker row (e.g. «РАБОТЫ котлован») — matches a block word AND
            # carries no numbers. «Накладные, командировочные …» (has qty/cost) is
            # a data row, NOT a marker.
            if _BLOCK_RE.match(name) and not has_metrics:
                block_default = _BLOCK_TYPE[_BLOCK_RE.match(name).group(1).lower()]
                continue

            if block_default is None:
                continue
            # Skip trailing footer lines («Составил:», phone, notes) — no numbers.
            if not has_metrics:
                continue

            item_type = classify_sectioned_row(name, None, _to_str(unit), block_default)
            unit_price = _to_optional_float(work_cost, treat_zero_as_none=True) \
                or _to_optional_float(mat_cost, treat_zero_as_none=True)

            results.append(ParsedRow(
                section=_BLOCK_LABEL.get(block_default, block_default),
                work_name=name,
                unit=_to_str(unit),
                quantity=_to_optional_float(qty),
                unit_price=unit_price,
                total_price=_to_optional_float(total, treat_zero_as_none=True),
                row_order=order,
                raw_data={
                    "item_type": item_type,
                    "block": block_default,
                    "parser_profile": self.name,
                    "source_strategy": self.name,
                },
                source_strategy=self.name,
            ))
            order += 1

            # Work rows: extract the machine they mention as a separate mechanism.
            if item_type == "work":
                mech_name = extract_mechanism_token(name)
                if mech_name:
                    results.append(ParsedRow(
                        section=_BLOCK_LABEL.get(block_default, block_default),
                        work_name=mech_name,
                        unit=None, quantity=None, unit_price=None, total_price=None,
                        row_order=order,
                        raw_data={
                            "item_type": "mechanism",
                            "source": "derived_from_work",
                            "linked_work": name,
                            "parser_profile": self.name,
                            "source_strategy": self.name,
                        },
                        source_strategy=self.name,
                    ))
                    order += 1

        return results, declared


_BLOCK_LABEL = {"work": "Работы", "material": "Материалы", "overhead": "Накладные"}
