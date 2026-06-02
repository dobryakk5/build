"""Parser profile ``excel_typed_journal`` — Excel estimates that carry an
explicit "Тип" column (Работа / Материал / Механизм / Накладные / Люди).

Unlike the generic ``StructuredSmetaParser`` type1 path (which folds materials
into their parent work and drops mechanism/overhead rows), this profile trusts
the "Тип" column and emits EVERY priceable row as its own typed line, so the
preview/Gantt split is driven by the operator's own classification.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .excel_parser import (
    StructuredSmetaParser,
    _cell_text,
    _to_optional_float,
    _to_str,
    is_subtotal_label,
)
from .pdf_parser import ParsedRow
from .resource_classifier import normalize_explicit_type

PROFILE_NAME = "excel_typed_journal"


class ExcelTypedJournalParser:
    name = PROFILE_NAME

    def parse(self, file_path: str | Path) -> tuple[list[ParsedRow], dict]:
        file_path = Path(file_path)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        smeta = StructuredSmetaParser()
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                info = smeta._find_type1_header(ws)
                if not info:
                    continue
                rows = self._parse(ws, info, smeta)
                return rows, {
                    "strategy": self.name,
                    "sheet": sheet_name,
                    "confidence": 1.0,
                    "rows_found": len(rows),
                }
        finally:
            wb.close()

        raise ValueError(
            "Профиль «Excel: колонка Тип» требует столбец «Тип» "
            "(Работа / Материал / Механизм / Накладные). Он не найден в файле."
        )

    def _parse(self, ws, info: dict, smeta: StructuredSmetaParser) -> list[ParsedRow]:
        results: list[ParsedRow] = []
        section: str | None = None
        order = 0

        for row_idx in range(info["header_row"] + 1, ws.max_row + 1):
            num = ws.cell(row_idx, info["num"]).value
            name = _cell_text(ws.cell(row_idx, info["name"]).value)
            tipo = _cell_text(ws.cell(row_idx, info["type"]).value)
            unit = ws.cell(row_idx, info["unit"]).value if info.get("unit") else None
            qty = ws.cell(row_idx, info["qty"]).value if info.get("qty") else None
            price = ws.cell(row_idx, info["price"]).value if info.get("price") else None
            total = ws.cell(row_idx, info["total"]).value if info.get("total") else None

            if not any((num, name, tipo, unit, qty, price, total)):
                continue

            num_str = _cell_text(num)
            if smeta._is_group_row_type1(ws, row_idx, info["name"], num_str, name, unit, qty):
                section = name or section
                continue

            if not name or is_subtotal_label(name):
                continue

            item_type, subtype = normalize_explicit_type(tipo)
            results.append(ParsedRow(
                section=section,
                work_name=name,
                unit=_to_str(unit),
                quantity=_to_optional_float(qty),
                unit_price=_to_optional_float(price, treat_zero_as_none=True),
                total_price=_to_optional_float(total, treat_zero_as_none=True),
                row_order=order,
                raw_data={
                    "num": num_str,
                    "type": tipo,
                    "item_type": item_type,
                    "resource_subtype": subtype,
                    "parser_profile": self.name,
                },
                source_strategy=self.name,
            ))
            order += 1

        return results
