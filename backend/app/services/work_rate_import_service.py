"""Excel importer for the work-rate catalogue.

Two source profiles are supported:
* normalized_rate_catalog_v1 — the five 11-column rate catalogues;
* market_estimate_observation_v1 — commercial estimates used as observations.

The importer is intentionally independent of the ORM.  It returns dataclasses
from :mod:`work_rate_models`; a repository/DB adapter can persist them.
"""
from __future__ import annotations

import ast
import hashlib
import json
import math
import re
from dataclasses import replace
from pathlib import Path
from typing import Any, Iterable
from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import range_boundaries

from app.services.work_rate_models import (
    LABOR_DERIVED,
    LABOR_UNKNOWN,
    MAPPING_STATUS_EXCLUDED,
    MAPPING_STATUS_OBSERVATION,
    MAPPING_STATUS_UNMAPPED,
    REVIEW_NEEDED,
    REVIEW_NEW,
    SOURCE_NORMALIZED,
    SOURCE_OBSERVATION,
    WorkRateImportResult,
    WorkRateImportRun,
    WorkRateItem,
    WorkRateSource,
    utcnow_iso,
)

PROFILE_NORMALIZED = "normalized_rate_catalog_v1"
PROFILE_OBSERVATION = "market_estimate_observation_v1"
DEFAULT_MAX_HEADER_SCAN_ROWS = 50
REFERENCE_MARKER_RE = re.compile(r"\[reference:\d+\]", re.IGNORECASE)
CELL_REF_RE = re.compile(r"\b([A-Z]{1,3}\d+)\b")
SUM_RE = re.compile(r"SUM\(([^()]+)\)", re.IGNORECASE)


UNIT_ALIASES: dict[str, tuple[str, str, float]] = {
    "м2": ("m2", "area", 1.0),
    "м²": ("m2", "area", 1.0),
    "кв м": ("m2", "area", 1.0),
    "кв.м": ("m2", "area", 1.0),
    "кв.м.": ("m2", "area", 1.0),
    "м3": ("m3", "volume", 1.0),
    "м³": ("m3", "volume", 1.0),
    "куб м": ("m3", "volume", 1.0),
    "куб.м": ("m3", "volume", 1.0),
    "куб.м.": ("m3", "volume", 1.0),
    "мп": ("m", "length", 1.0),
    "м.п": ("m", "length", 1.0),
    "м.п.": ("m", "length", 1.0),
    "пог м": ("m", "length", 1.0),
    "пог.м": ("m", "length", 1.0),
    "м": ("m", "length", 1.0),
    "мм": ("mm", "length", 1.0),
    "шт": ("pcs", "count", 1.0),
    "шт.": ("pcs", "count", 1.0),
    "т": ("t", "weight", 1.0),
    "кг": ("kg", "weight", 1.0),
    "чел час": ("person_hour", "time", 1.0),
    "чел-час": ("person_hour", "time", 1.0),
    "чел.-час": ("person_hour", "time", 1.0),
    "чел ч": ("person_hour", "time", 1.0),
    "чел.-ч": ("person_hour", "time", 1.0),
    "маш час": ("machine_hour", "machine_time", 1.0),
    "маш-час": ("machine_hour", "machine_time", 1.0),
    "маш.-час": ("machine_hour", "machine_time", 1.0),
    "маш ч": ("machine_hour", "machine_time", 1.0),
    "маш.-ч": ("machine_hour", "machine_time", 1.0),
    "сотка": ("are", "area_plot", 1.0),
    "сотки": ("are", "area_plot", 1.0),
    "соток": ("are", "area_plot", 1.0),
    "смена": ("shift", "time_scope", 1.0),
    "смен": ("shift", "time_scope", 1.0),
    "смены": ("shift", "time_scope", 1.0),
    "компл": ("set", "scope", 1.0),
    "компл.": ("set", "scope", 1.0),
    "комплект": ("set", "scope", 1.0),
    "точка": ("point", "count_scope", 1.0),
    "точки": ("point", "count_scope", 1.0),
    "проем": ("opening", "count_scope", 1.0),
    "проём": ("opening", "count_scope", 1.0),
    "проема": ("opening", "count_scope", 1.0),
    "проёма": ("opening", "count_scope", 1.0),
    "участок": ("site", "scope", 1.0),
    "участка": ("site", "scope", 1.0),
    "окно": ("window", "count_scope", 1.0),
    "окон": ("window", "count_scope", 1.0),
    "окна": ("window", "count_scope", 1.0),
    "%": ("percent", "ratio", 1.0),
}

HEADER_TOKEN_GROUPS: dict[str, list[tuple[str, ...]]] = {
    PROFILE_NORMALIZED: [
        ("вид", "работ"),
        ("ед", "изм"),
        ("расценк",),
    ],
    PROFILE_OBSERVATION: [
        ("наименован",),
        ("ед", "изм"),
        ("кол",),
    ],
}

OVERHEAD_RE = re.compile(
    r"\b(?:накладн|командиров|непредвиденн|резерв|расходн(?:ые)?\s+материал|бытовк|"
    r"погрузо\s*/?\s*разгрузочн|аренд[аы]\s+спецтехник|вывоз\s+мусора)\b",
    re.IGNORECASE,
)
LOGISTICS_RE = re.compile(r"\b(?:доставк|разгрузк|погрузк|складирован|перенос\s+материал)\b", re.I)


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_header(value: Any) -> str:
    text = str(value or "").casefold().replace("ё", "е")
    text = re.sub(r"[,\.()\-/\\]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_reference_markers(value: Any) -> str:
    text = str(value or "")
    text = REFERENCE_MARKER_RE.sub("", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_name(value: Any) -> str:
    text = clean_reference_markers(value).casefold().replace("ё", "е")
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ;")


def normalize_unit(value: Any) -> tuple[str | None, str | None, float | None]:
    if value is None:
        return None, None, None
    raw = normalize_header(value)
    raw = raw.replace("кв м", "кв м").strip()
    found = UNIT_ALIASES.get(raw)
    if found:
        return found
    compact = raw.replace(" ", "")
    found = UNIT_ALIASES.get(compact)
    if found:
        return found
    return None, None, None


def normalize_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "—", "–"}:
        return None
    text = text.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    text = re.sub(r"(?i)(руб\.?|р\.?|₽|чел\.?-?ч|%)$", "", text).strip()
    try:
        return float(text)
    except ValueError:
        return None


def _row_text(values: Iterable[Any]) -> str:
    return normalize_header(" ".join(str(v) for v in values if v is not None))


def find_header_row(ws, profile: str, max_scan_rows: int = DEFAULT_MAX_HEADER_SCAN_ROWS) -> int | None:
    groups = HEADER_TOKEN_GROUPS[profile]
    for index, row in enumerate(
        ws.iter_rows(max_row=min(ws.max_row, max_scan_rows), values_only=True), 1
    ):
        text = _row_text(row)
        if all(all(token in text for token in group) for group in groups):
            return index
    return None


def _contains(normalized: str, *tokens: str) -> bool:
    return all(token in normalized for token in tokens)


def map_normalized_columns(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[header_row]:
        header = normalize_header(cell.value)
        col = int(cell.column)
        if not header:
            continue
        if header in {"№", "n"} or header.startswith("№"):
            mapping.setdefault("external_code", col)
        elif _contains(header, "вид", "работ"):
            mapping.setdefault("name", col)
        elif _contains(header, "ед", "изм"):
            mapping.setdefault("unit", col)
        elif _contains(header, "расценк", "мин"):
            mapping.setdefault("price_min", col)
        elif _contains(header, "расценк", "макс"):
            mapping.setdefault("price_max", col)
        elif _contains(header, "расценк", "сред"):
            mapping.setdefault("price_avg", col)
        elif ("трудоемк" in header or "трудоёмк" in header) and "мин" in header:
            mapping.setdefault("labor_min", col)
        elif ("трудоемк" in header or "трудоёмк" in header) and "макс" in header:
            mapping.setdefault("labor_max", col)
        elif ("трудоемк" in header or "трудоёмк" in header) and "сред" in header:
            mapping.setdefault("labor_avg", col)
        elif _contains(header, "часов", "ставк"):
            mapping.setdefault("hourly_rate", col)
        elif "примеч" in header:
            mapping.setdefault("notes", col)
    required = {"name", "unit", "price_min", "price_max", "price_avg"}
    missing = required - set(mapping)
    if missing:
        raise ValueError("normalized catalogue headers missing: " + ", ".join(sorted(missing)))
    return mapping


def map_observation_columns(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[header_row]:
        header = normalize_header(cell.value)
        col = int(cell.column)
        if header.startswith("№"):
            mapping.setdefault("external_code", col)
        elif "наименован" in header:
            mapping.setdefault("name", col)
        elif _contains(header, "ед", "изм"):
            mapping.setdefault("unit", col)
        elif "кол" in header:
            mapping.setdefault("quantity", col)
        elif _contains(header, "стоимост", "единиц"):
            mapping.setdefault("price_avg", col)
        elif header == "всего" or header.startswith("всего "):
            mapping.setdefault("total_price", col)
    # Commercial file uses a merged two-level header.  Known positions are a
    # safe fallback after the header row itself was found semantically.
    mapping.setdefault("external_code", 1)
    mapping.setdefault("name", 2)
    mapping.setdefault("unit", 3)
    mapping.setdefault("quantity", 4)
    mapping.setdefault("price_avg", 5)
    mapping.setdefault("total_price", 7)
    return mapping


class SafeFormulaEvaluator:
    """Small evaluator for arithmetic, cell references and SUM ranges."""

    def __init__(self, formula_ws, value_ws):
        self.formula_ws = formula_ws
        self.value_ws = value_ws
        self._stack: set[str] = set()

    def cell_value(self, coordinate: str) -> float | None:
        coordinate = coordinate.upper()
        cached = self.value_ws[coordinate].value
        number = normalize_number(cached)
        if number is not None:
            return number
        raw = self.formula_ws[coordinate].value
        if isinstance(raw, str) and raw.startswith("="):
            if coordinate in self._stack:
                return None
            self._stack.add(coordinate)
            try:
                return self.evaluate(raw)
            finally:
                self._stack.discard(coordinate)
        return normalize_number(raw)

    def _replace_sum(self, expression: str) -> str:
        while True:
            match = SUM_RE.search(expression)
            if not match:
                return expression
            token = match.group(1).strip()
            total = 0.0
            if ":" in token:
                min_col, min_row, max_col, max_row = range_boundaries(token)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        value = self.cell_value(self.formula_ws.cell(row, col).coordinate)
                        total += float(value or 0)
            else:
                for coordinate in token.split(","):
                    total += float(self.cell_value(coordinate.strip()) or 0)
            expression = expression[: match.start()] + str(total) + expression[match.end() :]

    def evaluate(self, formula: str) -> float | None:
        expression = formula.strip()
        if expression.startswith("="):
            expression = expression[1:]
        expression = self._replace_sum(expression)

        def repl(match: re.Match[str]) -> str:
            value = self.cell_value(match.group(1))
            return "0" if value is None else repr(float(value))

        expression = CELL_REF_RE.sub(repl, expression)
        try:
            node = ast.parse(expression, mode="eval")
            return float(self._eval_node(node.body))
        except (SyntaxError, TypeError, ValueError, ZeroDivisionError, OverflowError):
            return None

    def _eval_node(self, node: ast.AST) -> float:
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            value = self._eval_node(node.operand)
            return value if isinstance(node.op, ast.UAdd) else -value
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
            left = self._eval_node(node.left)
            right = self._eval_node(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            return left / right
        raise ValueError("unsupported formula expression")


def _cell_numeric(
    formula_cell: Cell,
    value_cell: Cell,
    evaluator: SafeFormulaEvaluator,
) -> tuple[float | None, dict[str, Any]]:
    formula_text = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
    cached_value = value_cell.value
    value = normalize_number(cached_value)
    evaluated = False
    if value is None and formula_text:
        value = evaluator.evaluate(formula_text)
        evaluated = value is not None
    return value, {
        "formula_text": formula_text,
        "cached_value": cached_value,
        "evaluated_by_safe_evaluator": evaluated,
    }


def _deterministic_source_id(source_kind: str, filename: str, sheet: str) -> str:
    return str(uuid5(NAMESPACE_URL, f"work-rate:{source_kind}:{filename}:{sheet}"))


def _stable_key(
    source_id: str,
    sheet: str,
    source_row: int,
    external_code: str | None,
) -> str:
    identity = external_code.strip() if external_code and external_code.strip() else f"row:{source_row}"
    return hashlib.sha256(f"{source_id}|{sheet}|{identity}".encode("utf-8")).hexdigest()


def _content_hash(item_payload: dict[str, Any]) -> str:
    return hashlib.sha256(
        json.dumps(item_payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def _fill_range_values(
    minimum: float | None,
    maximum: float | None,
    average: float | None,
) -> tuple[float | None, float | None, float | None]:
    values = [v for v in (minimum, maximum, average) if v is not None]
    if not values:
        return None, None, None
    if average is None and minimum is not None and maximum is not None:
        average = (minimum + maximum) / 2
    elif average is None:
        average = minimum if minimum is not None else maximum
    if minimum is None:
        minimum = average
    if maximum is None:
        maximum = average
    return minimum, maximum, average


def _validate_range(minimum: float | None, average: float | None, maximum: float | None) -> bool:
    values = [v for v in (minimum, average, maximum) if v is not None]
    if any(v < 0 for v in values):
        return False
    if minimum is not None and average is not None and minimum > average:
        return False
    if average is not None and maximum is not None and average > maximum:
        return False
    return True


def _classify_observation_role(name: str) -> tuple[str, str]:
    normalized = normalize_name(name)
    # Standalone waste/soil removal is a schedulable domain operation unless
    # the commercial estimate explicitly marks it as a customer-provided item.
    if any(marker in normalized for marker in (
        "утилизация грунта",
        "вывоз грунта",
        "вывоз строительного мусора",
    )) and "заказчик" not in normalized:
        return "work", MAPPING_STATUS_OBSERVATION
    if any(marker in normalized for marker in (
        "расходные материалы",
        "накладные",
        "командировочные",
        "непредвиденные",
        "резерв",
        "бытовка",
    )):
        return "overhead", MAPPING_STATUS_EXCLUDED
    if any(marker in normalized for marker in (
        "погрузо/разгрузочные",
        "погрузо разгрузочные",
        "аренда спецтехники",
        "вывоз мусора",
        "доставка",
        "разгрузка",
        "завоз и раскладка материалов",
    )):
        return "logistics", MAPPING_STATUS_EXCLUDED
    return "work", MAPPING_STATUS_OBSERVATION


class WorkRateImportService:
    def __init__(self, *, max_header_scan_rows: int = DEFAULT_MAX_HEADER_SCAN_ROWS):
        self.max_header_scan_rows = max_header_scan_rows

    def detect_profile(self, ws) -> tuple[str, int]:
        normalized_row = find_header_row(ws, PROFILE_NORMALIZED, self.max_header_scan_rows)
        if normalized_row is not None:
            return PROFILE_NORMALIZED, normalized_row
        observation_row = find_header_row(ws, PROFILE_OBSERVATION, self.max_header_scan_rows)
        if observation_row is not None:
            return PROFILE_OBSERVATION, observation_row
        raise ValueError("header_not_found")

    def import_file(
        self,
        path: str | Path,
        *,
        previous_items: Iterable[WorkRateItem] | None = None,
        source_version: str = "1",
    ) -> WorkRateImportResult:
        path = Path(path)
        file_hash = file_sha256(path)
        formula_wb = load_workbook(path, data_only=False, read_only=False)
        value_wb = load_workbook(path, data_only=True, read_only=False)
        previous_by_key = {
            item.stable_row_key: item
            for item in (previous_items or [])
            if item.stable_row_key
        }

        all_items: list[WorkRateItem] = []
        source: WorkRateSource | None = None
        run = WorkRateImportRun(filename=path.name, file_hash=file_hash, status="running")

        for sheet_name in formula_wb.sheetnames:
            formula_ws = formula_wb[sheet_name]
            value_ws = value_wb[sheet_name]
            try:
                profile, header_row = self.detect_profile(value_ws)
            except ValueError:
                run.errors_json.append({"sheet": sheet_name, "error": "header_not_found"})
                continue

            source_kind = SOURCE_NORMALIZED if profile == PROFILE_NORMALIZED else SOURCE_OBSERVATION
            source_id = _deterministic_source_id(source_kind, path.name, sheet_name)
            if source is None:
                source = WorkRateSource(
                    id=source_id,
                    name=path.stem,
                    source_kind=source_kind,
                    source_file=path.name,
                    source_sheet=sheet_name,
                    source_version=source_version,
                    hourly_rate=800.0 if source_kind == SOURCE_NORMALIZED else None,
                    labor_basis=LABOR_DERIVED if source_kind == SOURCE_NORMALIZED else LABOR_UNKNOWN,
                    metadata_json={"profile": profile, "file_hash": file_hash},
                )
                run.source_id = source_id

            evaluator = SafeFormulaEvaluator(formula_ws, value_ws)
            if profile == PROFILE_NORMALIZED:
                columns = map_normalized_columns(value_ws, header_row)
                items = self._parse_normalized_sheet(
                    formula_ws,
                    value_ws,
                    evaluator,
                    columns,
                    header_row,
                    source,
                    previous_by_key,
                )
            else:
                columns = map_observation_columns(value_ws, header_row)
                items = self._parse_observation_sheet(
                    formula_ws,
                    value_ws,
                    evaluator,
                    columns,
                    header_row,
                    source,
                    previous_by_key,
                )
            all_items.extend(items)

        if source is None:
            source = WorkRateSource(
                id=_deterministic_source_id(SOURCE_NORMALIZED, path.name, "unknown"),
                name=path.stem,
                source_file=path.name,
                source_kind=SOURCE_NORMALIZED,
                metadata_json={"file_hash": file_hash},
            )
            run.source_id = source.id
            run.status = "failed"
        else:
            run.status = "completed" if not run.errors_json else "completed_with_warnings"

        run.rows_total = len(all_items)
        run.rows_imported = len(all_items)
        run.rows_created = sum(1 for item in all_items if item.revision == 1)
        run.rows_updated = sum(1 for item in all_items if item.revision > 1)
        run.rows_unmapped = sum(1 for item in all_items if item.mapping_status == MAPPING_STATUS_UNMAPPED)
        run.rows_needs_review = sum(1 for item in all_items if item.review_status == REVIEW_NEEDED)
        run.finished_at = utcnow_iso()
        return WorkRateImportResult(source=source, run=run, items=all_items)

    def _build_item(
        self,
        *,
        source: WorkRateSource,
        sheet: str,
        row_index: int,
        external_code: str | None,
        name: str,
        notes: str | None,
        unit_raw: str | None,
        quantity: float | None,
        price_min: float | None,
        price_max: float | None,
        price_avg: float | None,
        total_price: float | None,
        labor_min: float | None,
        labor_max: float | None,
        labor_avg: float | None,
        hourly_rate: float | None,
        mapping_status: str,
        row_role: str,
        source_payload: dict[str, Any],
        previous_by_key: dict[str, WorkRateItem],
    ) -> WorkRateItem:
        normalized = normalize_name(name)
        normalized_notes = normalize_name(notes or "") or None
        unit_code, unit_dimension, _ = normalize_unit(unit_raw)
        stable_key = _stable_key(source.id, sheet, row_index, external_code)

        price_min, price_max, price_avg = _fill_range_values(price_min, price_max, price_avg)
        labor_min, labor_max, labor_avg = _fill_range_values(labor_min, labor_max, labor_avg)
        payload_for_hash = {
            "normalized_name": normalized,
            "unit_code": unit_code,
            "price_min": price_min,
            "price_max": price_max,
            "price_avg": price_avg,
            "labor_min": labor_min,
            "labor_max": labor_max,
            "labor_avg": labor_avg,
            "hourly_rate": hourly_rate,
            "normalized_notes": normalized_notes,
            "quantity": quantity,
            "total_price": total_price,
        }
        content_hash = _content_hash(payload_for_hash)
        previous = previous_by_key.get(stable_key)
        revision = 1
        supersedes = None
        review_status = REVIEW_NEW
        review_reason = None
        if previous is not None:
            if previous.row_content_hash == content_hash:
                # Reuse stable identity for unchanged content.
                return replace(previous, source_payload=source_payload, updated_at=utcnow_iso())
            revision = previous.revision + 1
            supersedes = previous.id
            review_status = REVIEW_NEEDED
            review_reason = "source_row_changed"

        if unit_code is None and unit_raw:
            review_status = REVIEW_NEEDED
            review_reason = review_reason or "unknown_unit"
        if not _validate_range(price_min, price_avg, price_max) or not _validate_range(labor_min, labor_avg, labor_max):
            review_status = REVIEW_NEEDED
            review_reason = review_reason or "invalid_numeric_range"

        if source.source_kind == SOURCE_NORMALIZED and hourly_rate and price_avg is not None and labor_avg is not None:
            expected = price_avg / hourly_rate
            tolerance = max(0.05, expected * 0.03)
            if abs(expected - labor_avg) > tolerance:
                review_status = REVIEW_NEEDED
                review_reason = review_reason or "derived_labor_mismatch"

        return WorkRateItem(
            source_id=source.id,
            source_row=row_index,
            external_code=external_code,
            stable_row_key=stable_key,
            row_content_hash=content_hash,
            revision=revision,
            supersedes_rate_item_id=supersedes,
            name=name.strip(),
            normalized_name=normalized,
            notes=notes.strip() if notes else None,
            normalized_notes=normalized_notes,
            unit_raw=unit_raw,
            unit_code=unit_code,
            unit_dimension=unit_dimension,
            quantity=quantity,
            price_min=price_min,
            price_max=price_max,
            price_avg=price_avg,
            total_price=total_price,
            labor_min=labor_min,
            labor_max=labor_max,
            labor_avg=labor_avg,
            hourly_rate=hourly_rate,
            labor_basis=LABOR_DERIVED if source.source_kind == SOURCE_NORMALIZED else LABOR_UNKNOWN,
            mapping_status=mapping_status,
            review_status=review_status,
            review_reason=review_reason,
            row_role=row_role,
            source_payload=source_payload,
        )

    def _parse_normalized_sheet(
        self,
        formula_ws,
        value_ws,
        evaluator: SafeFormulaEvaluator,
        columns: dict[str, int],
        header_row: int,
        source: WorkRateSource,
        previous_by_key: dict[str, WorkRateItem],
    ) -> list[WorkRateItem]:
        items: list[WorkRateItem] = []
        for row_index in range(header_row + 1, value_ws.max_row + 1):
            name = value_ws.cell(row_index, columns["name"]).value
            if not name or not str(name).strip():
                continue
            external = value_ws.cell(row_index, columns.get("external_code", 1)).value
            unit_raw = value_ws.cell(row_index, columns["unit"]).value
            notes = value_ws.cell(row_index, columns.get("notes", value_ws.max_column)).value if columns.get("notes") else None
            numeric: dict[str, float | None] = {}
            formulas: dict[str, Any] = {}
            for key in (
                "price_min", "price_max", "price_avg", "labor_min", "labor_max", "labor_avg", "hourly_rate"
            ):
                col = columns.get(key)
                if not col:
                    numeric[key] = None
                    continue
                value, diagnostics = _cell_numeric(
                    formula_ws.cell(row_index, col),
                    value_ws.cell(row_index, col),
                    evaluator,
                )
                numeric[key] = value
                formulas[key] = diagnostics
            source_payload = {
                "sheet": value_ws.title,
                "source_row": row_index,
                "raw_name": name,
                "raw_notes": notes,
                "raw_unit": unit_raw,
                "formula_diagnostics": formulas,
            }
            normalized_unit_code, _, _ = normalize_unit(unit_raw)
            row_role, role_status = _classify_observation_role(str(name))
            if normalized_unit_code == "percent":
                row_role, role_status = "overhead", MAPPING_STATUS_EXCLUDED
            normalized_status = (
                MAPPING_STATUS_EXCLUDED
                if role_status == MAPPING_STATUS_EXCLUDED
                else MAPPING_STATUS_UNMAPPED
            )
            item = self._build_item(
                source=source,
                sheet=value_ws.title,
                row_index=row_index,
                external_code=str(external).strip() if external is not None else None,
                name=str(name),
                notes=str(notes) if notes is not None else None,
                unit_raw=str(unit_raw).strip() if unit_raw is not None else None,
                quantity=None,
                price_min=numeric["price_min"],
                price_max=numeric["price_max"],
                price_avg=numeric["price_avg"],
                total_price=None,
                labor_min=numeric["labor_min"],
                labor_max=numeric["labor_max"],
                labor_avg=numeric["labor_avg"],
                hourly_rate=numeric["hourly_rate"] or source.hourly_rate,
                mapping_status=normalized_status,
                row_role=row_role,
                source_payload=source_payload,
                previous_by_key=previous_by_key,
            )
            items.append(item)
        return items

    def _parse_observation_sheet(
        self,
        formula_ws,
        value_ws,
        evaluator: SafeFormulaEvaluator,
        columns: dict[str, int],
        header_row: int,
        source: WorkRateSource,
        previous_by_key: dict[str, WorkRateItem],
    ) -> list[WorkRateItem]:
        items: list[WorkRateItem] = []
        for row_index in range(header_row + 1, value_ws.max_row + 1):
            external = value_ws.cell(row_index, columns["external_code"]).value
            name = value_ws.cell(row_index, columns["name"]).value
            # Section labels and totals have no numeric item code.
            if external is None or name is None:
                continue
            external_text = str(external).strip()
            if not re.fullmatch(r"\d+(?:\.0+)?", external_text):
                continue
            name_text = str(name).strip()
            if not name_text:
                continue
            unit_raw = value_ws.cell(row_index, columns["unit"]).value
            # Multi-row headers often contain the helper row 1/2/3/4/5/6/7.
            if name_text.isdigit() and str(unit_raw or "").strip().isdigit():
                continue
            quantity = normalize_number(value_ws.cell(row_index, columns["quantity"]).value)
            price, price_diag = _cell_numeric(
                formula_ws.cell(row_index, columns["price_avg"]),
                value_ws.cell(row_index, columns["price_avg"]),
                evaluator,
            )
            total, total_diag = _cell_numeric(
                formula_ws.cell(row_index, columns["total_price"]),
                value_ws.cell(row_index, columns["total_price"]),
                evaluator,
            )
            row_role, mapping_status = _classify_observation_role(name_text)
            source_payload = {
                "sheet": value_ws.title,
                "source_row": row_index,
                "raw_name": name,
                "raw_unit": unit_raw,
                "formula_diagnostics": {
                    "price_avg": price_diag,
                    "total_price": total_diag,
                },
            }
            item = self._build_item(
                source=source,
                sheet=value_ws.title,
                row_index=row_index,
                external_code=external_text,
                name=name_text,
                notes=None,
                unit_raw=str(unit_raw).strip() if unit_raw is not None else None,
                quantity=quantity,
                price_min=price,
                price_max=price,
                price_avg=price,
                total_price=total,
                labor_min=None,
                labor_max=None,
                labor_avg=None,
                hourly_rate=None,
                mapping_status=mapping_status,
                row_role=row_role,
                source_payload=source_payload,
                previous_by_key=previous_by_key,
            )
            if mapping_status == MAPPING_STATUS_EXCLUDED:
                item.review_status = REVIEW_NEW
                item.auto_applicable = False
            elif item.unit_code == "percent":
                item.mapping_status = MAPPING_STATUS_EXCLUDED
                item.row_role = "overhead"
                item.auto_applicable = False
            items.append(item)
        return items
