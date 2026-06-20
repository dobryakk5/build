"""Excel profile for estimates where work rows are numbered ``N`` and their
materials/equipment are numbered ``N.M``.

The parser uses the structural number as the source of truth.  Text
classification must not turn nested equipment into independent works.
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Optional

import openpyxl

from .excel_parser import _cell_text, _to_optional_float, _to_str
from .pdf_parser import ParsedRow

PROFILE_NAME = "excel_work_material_matrix"
MIN_DETECT_CONFIDENCE = 0.95

_WORK_NUM_RE = re.compile(r"^\d+$")
_MATERIAL_NUM_RE = re.compile(r"^(\d+)\.(\d+)$")
_TOTAL_WITHOUT_VAT_RE = re.compile(r"^\s*итого\s*$", re.IGNORECASE)
_VAT_RE = re.compile(r"^\s*ндс(?:\s+(\d+(?:[.,]\d+)?))?\s*%?\s*$", re.IGNORECASE)
_GRAND_TOTAL_RE = re.compile(r"^\s*всего\s+по\s+смете\s*$", re.IGNORECASE)
_LETTER_RE = re.compile(r"[A-Za-zА-Яа-яЁё]")
_MONEY_QUANT = Decimal("0.01")


def _norm_text(value: Any) -> str:
    return re.sub(r"\s+", " ", _cell_text(value).replace("\xa0", " ")).strip()


def _norm_header(value: Any) -> str:
    return _norm_text(value).lower().replace("ё", "е")


def _normalize_position_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        raw = format(value, "f").rstrip("0").rstrip(".")
    else:
        raw = re.sub(r"\s+", "", str(value).strip()).replace(",", ".")

    raw = raw.strip("'")
    if not re.fullmatch(r"\d+(?:\.\d+)?", raw):
        return raw
    return ".".join(str(int(part)) for part in raw.split("."))


def _has_value(value: Optional[float]) -> bool:
    return value is not None and value != 0


def _is_empty_cost(value: Optional[float]) -> bool:
    return value is None or value == 0


def _money(value: Any) -> Optional[float]:
    """Return an Excel money value rounded to kopecks with decimal arithmetic."""
    if value is None:
        return None
    try:
        return float(Decimal(str(value)).quantize(_MONEY_QUANT, rounding=ROUND_HALF_UP))
    except (InvalidOperation, TypeError, ValueError):
        return None


class ExcelWorkMaterialMatrixParser:
    """Parse multi-row cost headers with work/material split columns."""

    name = PROFILE_NAME

    def detect(self, file_path: str | Path) -> dict:
        path = Path(file_path)
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        except Exception as exc:
            return {
                "detected": False,
                "confidence": 0.0,
                "strategy": self.name,
                "reason": f"workbook_open_error:{type(exc).__name__}",
            }
        try:
            return self._detect_workbook(wb)
        finally:
            wb.close()

    def parse(self, file_path: str | Path) -> tuple[list[ParsedRow], dict]:
        path = Path(file_path)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        try:
            detected = self._detect_workbook(wb)
            if detected.get("confidence", 0.0) < MIN_DETECT_CONFIDENCE:
                raise ValueError(
                    "Профиль «Excel: работы с материалами по подномерам» не нашёл "
                    "структуру N / N.M с раздельными колонками стоимости работ и материалов."
                )
            ws = wb[detected["sheet"]]
            rows, meta = self._parse_sheet(ws, detected["column_map"])
            meta.update({
                "strategy": self.name,
                "parser_profile": self.name,
                "sheet": ws.title,
                "confidence": detected["confidence"],
                "detector": detected.get("stats", {}),
            })
            return rows, meta
        finally:
            wb.close()

    def _detect_workbook(self, wb) -> dict:
        best: dict = {
            "detected": False,
            "confidence": 0.0,
            "strategy": self.name,
            "sheet": None,
            "column_map": None,
            "stats": {},
        }
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            column_map = self._find_column_map(ws)
            if not column_map:
                continue
            stats = self._analyze_structure(ws, column_map)
            confidence = self._confidence(stats)
            if confidence > best["confidence"]:
                best = {
                    "detected": confidence >= MIN_DETECT_CONFIDENCE,
                    "confidence": confidence,
                    "strategy": self.name,
                    "sheet": sheet_name,
                    "column_map": column_map,
                    "stats": stats,
                }
        return best

    def _find_column_map(self, ws) -> Optional[dict[str, int]]:
        max_row = min(ws.max_row, 50)
        max_col = min(ws.max_column, 30)

        for row_idx in range(1, max_row + 1):
            texts = {c: _norm_header(ws.cell(row_idx, c).value) for c in range(1, max_col + 1)}
            num_col = next((c for c, t in texts.items() if t.startswith("№") or t in {"номер", "n пп"}), None)
            name_col = next((c for c, t in texts.items() if "наименование" in t), None)
            unit_col = next((c for c, t in texts.items() if "ед" in t and "изм" in t), None)
            qty_col = next((c for c, t in texts.items() if t.startswith("кол") or "количество" in t), None)
            unit_group_col = next((c for c, t in texts.items() if "стоимость единицы" in t), None)
            total_group_col = next((c for c, t in texts.items() if "общая стоимость" in t), None)

            if not all((num_col, name_col, unit_col, qty_col, unit_group_col, total_group_col)):
                continue
            if unit_group_col >= total_group_col:
                continue

            scan_end = min(ws.max_row, row_idx + 5)
            unit_work_col = unit_material_col = None
            total_work_col = total_material_col = None
            header_end = row_idx

            for r in range(row_idx + 1, scan_end + 1):
                for c in range(unit_group_col, total_group_col):
                    text = _norm_header(ws.cell(r, c).value)
                    if text.startswith("работ"):
                        unit_work_col = c
                        header_end = max(header_end, r)
                    elif text.startswith("материал"):
                        unit_material_col = c
                        header_end = max(header_end, r)
                for c in range(total_group_col, max_col + 1):
                    text = _norm_header(ws.cell(r, c).value)
                    if text.startswith("работ"):
                        total_work_col = c
                        header_end = max(header_end, r)
                    elif text.startswith("материал"):
                        total_material_col = c
                        header_end = max(header_end, r)

            if not all((unit_work_col, unit_material_col, total_work_col, total_material_col)):
                continue

            return {
                "header_start_row": row_idx,
                "header_end_row": header_end,
                "num": num_col,
                "name": name_col,
                "unit": unit_col,
                "qty": qty_col,
                "unit_total": unit_group_col,
                "unit_work": unit_work_col,
                "unit_material": unit_material_col,
                "total_total": total_group_col,
                "total_work": total_work_col,
                "total_material": total_material_col,
            }
        return None

    def _analyze_structure(self, ws, cols: dict[str, int]) -> dict:
        work_numbers: set[str] = set()
        material_parents: list[str] = []
        work_split_ok = 0
        material_split_ok = 0
        work_rows = 0
        material_rows = 0
        technical_rows = 0

        for row_idx in range(cols["header_end_row"] + 1, ws.max_row + 1):
            if self._is_technical_column_number_row(ws, row_idx, cols):
                technical_rows += 1
                continue

            num = _normalize_position_number(ws.cell(row_idx, cols["num"]).value)
            name = _norm_text(ws.cell(row_idx, cols["name"]).value)
            if not name:
                continue

            work_unit = _to_optional_float(ws.cell(row_idx, cols["unit_work"]).value)
            work_total = _to_optional_float(ws.cell(row_idx, cols["total_work"]).value)
            material_unit = _to_optional_float(ws.cell(row_idx, cols["unit_material"]).value)
            material_total = _to_optional_float(ws.cell(row_idx, cols["total_material"]).value)

            if _WORK_NUM_RE.fullmatch(num):
                work_rows += 1
                work_numbers.add(num)
                if (
                    (_has_value(work_unit) or _has_value(work_total))
                    and _is_empty_cost(material_unit)
                    and _is_empty_cost(material_total)
                ):
                    work_split_ok += 1
            else:
                material_match = _MATERIAL_NUM_RE.fullmatch(num)
                if material_match:
                    material_rows += 1
                    material_parents.append(str(int(material_match.group(1))))
                    if (
                        (_has_value(material_unit) or _has_value(material_total))
                        and _is_empty_cost(work_unit)
                        and _is_empty_cost(work_total)
                    ):
                        material_split_ok += 1

        linked_pairs = sum(1 for parent in material_parents if parent in work_numbers)
        return {
            "work_rows": work_rows,
            "material_rows": material_rows,
            "linked_pairs": linked_pairs,
            "work_split_ratio": work_split_ok / work_rows if work_rows else 0.0,
            "material_split_ratio": material_split_ok / material_rows if material_rows else 0.0,
            "technical_rows": technical_rows,
        }

    @staticmethod
    def _confidence(stats: dict) -> float:
        if stats.get("work_rows", 0) < 3 or stats.get("material_rows", 0) < 3:
            return 0.0
        if stats.get("linked_pairs", 0) < 3:
            return 0.0
        if stats.get("work_split_ratio", 0.0) < 0.8:
            return 0.0
        if stats.get("material_split_ratio", 0.0) < 0.8:
            return 0.0
        return 1.0

    def _parse_sheet(self, ws, cols: dict[str, int]) -> tuple[list[ParsedRow], dict]:
        works: list[ParsedRow] = []
        work_by_num: dict[str, ParsedRow] = {}
        materials: list[tuple[str, dict]] = []
        orphan_material_rows: list[dict] = []
        skipped_rows: list[dict] = []
        declared_totals: list[dict] = []
        current_section: Optional[str] = None
        work_order = 0
        skipped_technical_rows = 0
        source_rows_found = 0

        for row_idx in range(cols["header_end_row"] + 1, ws.max_row + 1):
            if self._is_technical_column_number_row(ws, row_idx, cols):
                skipped_technical_rows += 1
                source_rows_found += 1
                skipped_rows.append({
                    "source_excel_row": row_idx,
                    "skip_reason": "technical_column_number_row",
                })
                continue

            raw_num = ws.cell(row_idx, cols["num"]).value
            num = _normalize_position_number(raw_num)
            name = _norm_text(ws.cell(row_idx, cols["name"]).value)
            label = _norm_text(name or raw_num)

            if self._append_declared_total(ws, row_idx, cols, label, declared_totals):
                continue

            metrics = self._row_metrics(ws, row_idx, cols)
            if not num and name and all(value is None or value == 0 for value in metrics.values()):
                current_section = name
                continue

            if _WORK_NUM_RE.fullmatch(num) and name:
                source_rows_found += 1
                row = ParsedRow(
                    section=current_section,
                    work_name=name,
                    unit=_to_str(ws.cell(row_idx, cols["unit"]).value),
                    quantity=_to_optional_float(ws.cell(row_idx, cols["qty"]).value),
                    unit_price=metrics["unit_work"],
                    total_price=metrics["total_work"],
                    materials=[],
                    row_order=work_order,
                    raw_data={
                        "item_type": "work",
                        "item_type_confidence": 1.0,
                        "source_num": num,
                        "source_excel_row": row_idx,
                        "source_work_unit_price": metrics["unit_work"],
                        "source_material_unit_price": metrics["unit_material"],
                        "source_grand_unit_price": metrics["unit_total"],
                        "source_work_total": metrics["total_work"],
                        "source_material_total": metrics["total_material"],
                        "source_grand_total": metrics["total_total"],
                        "parser_profile": self.name,
                        "source_strategy": self.name,
                        **({"group_path": [current_section]} if current_section else {}),
                    },
                    source_strategy=self.name,
                )
                works.append(row)
                work_by_num.setdefault(num, row)
                work_order += 1
                continue

            material_match = _MATERIAL_NUM_RE.fullmatch(num)
            if material_match and name:
                source_rows_found += 1
                parent_num = str(int(material_match.group(1)))
                material = {
                    "name": name,
                    "unit": _to_str(ws.cell(row_idx, cols["unit"]).value),
                    "quantity": _to_optional_float(ws.cell(row_idx, cols["qty"]).value),
                    "unit_price": metrics["unit_material"],
                    "total_price": metrics["total_material"],
                    "source_num": num,
                    "parent_work_num": parent_num,
                    "source_excel_row": row_idx,
                    "item_type": "material",
                    "item_type_confidence": 1.0,
                    "source_work_unit_price": metrics["unit_work"],
                    "source_material_unit_price": metrics["unit_material"],
                    "source_grand_unit_price": metrics["unit_total"],
                    "source_work_total": metrics["total_work"],
                    "source_material_total": metrics["total_material"],
                    "source_grand_total": metrics["total_total"],
                }
                materials.append((parent_num, material))

        for parent_num, material in materials:
            parent = work_by_num.get(parent_num)
            if parent is not None:
                parent.materials.append(material)
            else:
                orphan_material_rows.append(material)

        return works, {
            "rows_found": len(works),
            "source_rows_found": source_rows_found,
            "work_rows_found": len(works),
            "material_rows_found": len(materials),
            "skipped_technical_rows": skipped_technical_rows,
            "skipped_rows": skipped_rows,
            "orphan_material_rows": orphan_material_rows,
            "declared_totals": declared_totals,
        }

    def _row_metrics(self, ws, row_idx: int, cols: dict[str, int]) -> dict[str, Optional[float]]:
        return {
            key: _to_optional_float(ws.cell(row_idx, cols[key]).value)
            for key in (
                "unit_total", "unit_work", "unit_material",
                "total_total", "total_work", "total_material",
            )
        }

    def _is_technical_column_number_row(self, ws, row_idx: int, cols: dict[str, int]) -> bool:
        values = [
            ws.cell(row_idx, cols[key]).value
            for key in (
                "num", "name", "unit", "qty", "unit_total", "unit_work",
                "unit_material", "total_total", "total_work", "total_material",
            )
        ]
        nonempty = [_norm_text(value) for value in values if _norm_text(value)]
        if len(nonempty) < 4:
            return False
        if any(_LETTER_RE.search(value) for value in nonempty):
            return False
        return all(re.fullmatch(r"[\d.,/\-]+", value) for value in nonempty)

    def _append_declared_total(
        self,
        ws,
        row_idx: int,
        cols: dict[str, int],
        label: str,
        declared_totals: list[dict],
    ) -> bool:
        if not label:
            return False
        total = _money(_to_optional_float(ws.cell(row_idx, cols["total_total"]).value))
        if _TOTAL_WITHOUT_VAT_RE.fullmatch(label):
            if total is not None:
                declared_totals.append({"kind": "total_without_vat", "label": label, "total": total})
            return True
        vat_match = _VAT_RE.fullmatch(label)
        if vat_match:
            if total is not None:
                item = {"kind": "vat", "label": label, "total": total}
                if vat_match.group(1):
                    item["rate"] = float(vat_match.group(1).replace(",", "."))
                declared_totals.append(item)
            return True
        if _GRAND_TOTAL_RE.fullmatch(label):
            if total is not None:
                declared_totals.append({"kind": "grand_total", "label": label, "total": total})
            return True
        return False
